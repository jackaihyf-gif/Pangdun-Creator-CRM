from __future__ import annotations

import json
import re
from collections import Counter, OrderedDict
from pathlib import Path
from typing import Any

import openpyxl


SOURCE = Path(r"C:\Users\18620\Downloads\铭瑄红人记者库.xlsx")
OUTPUT_DIR = Path("outputs/cleaning")
OUTPUT_JSON = OUTPUT_DIR / "cleaned_kol_data.json"

EXPECTED_HEADERS = [
    "类目",
    "国家",
    "母公司",
    "名字",
    "流量/粉丝",
    "网站类型",
    "链接",
    "报价",
    "合作情况",
    "合作链接",
    "联系人&职位",
    "联系方式",
    "合作备注",
    "合作备注2",
    "是否发产品Brief",
    "产品Brief 邮箱",
    "Press release邮箱",
]

STAGE_RULES = [
    ("Published", ["发布", "已发布"]),
    ("Brief Sent", ["已发brief", "已发产品brief", "发产品brief", "brief"]),
    ("In Production", ["测评", "评测", "制作中", "制作"]),
    ("Quoting", ["要钱", "报价", "收费", "付费"]),
    ("Waiting Reply", ["等回复", "待回复", "回复不活跃", "不活跃", "未回复"]),
    ("Contacted", ["已联系", "联系", "3 logic联系"]),
    ("To Contact", ["待开发", "开发"]),
]

EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+(?:\.[\w-]+)+")
PHONE_RE = re.compile(r"(?:\+?\d[\d\s().-]{6,}\d)")
URL_RE = re.compile(r"https?://[^\s)）]+", re.I)
PRODUCT_RE = re.compile(
    r"(?i)\b(?:"
    r"DDR[45]"
    r"|B\d{3}[A-Z0-9]*(?:\s+[A-Z][A-Z0-9-]*){0,3}"
    r"|X\d{3}[A-Z0-9]*(?:\s+[A-Z][A-Z0-9-]*){0,3}"
    r"|Z\d{3}[A-Z0-9]*(?:\s+[A-Z][A-Z0-9-]*){0,3}"
    r"|H\d{3}[A-Z0-9]*(?:\s+[A-Z][A-Z0-9-]*){0,3}"
    r"|AIGA"
    r"|iCraft"
    r"|MS-[A-Z0-9-]+"
    r")\b"
)


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_rows() -> tuple[list[str], list[dict[str, Any]], list[list[Any]]]:
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    headers = [text(v) for v in next(iterator)]
    usable_headers = headers[: len(EXPECTED_HEADERS)]
    raw_matrix = [usable_headers]
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(iterator, start=2):
        row_values = list(values[: len(EXPECTED_HEADERS)])
        if not any(v not in (None, "") for v in row_values):
            continue
        raw_matrix.append(row_values)
        item = {"raw_row": row_number}
        for index, header in enumerate(usable_headers):
            item[header] = row_values[index] if index < len(row_values) else None
        rows.append(item)
    return usable_headers, rows, raw_matrix


def normalize_media_key(row: dict[str, Any]) -> str:
    name = text(row.get("名字")).lower()
    link = text(row.get("链接")).lower()
    country = text(row.get("国家")).lower()
    if link:
        return f"name_link::{name}::{link}"
    return f"name_country::{name}::{country}"


def split_contact(value: str) -> tuple[str, str]:
    if not value:
        return "", ""
    separators = ["｜", "|", " - ", "，", ",", "：", ":"]
    for sep in separators:
        if sep in value:
            left, right = value.split(sep, 1)
            return left.strip(), right.strip()
    parts = value.split()
    if len(parts) >= 3:
        return " ".join(parts[:-1]), parts[-1]
    return value.strip(), ""


def infer_stage(row: dict[str, Any], notes_blob: str) -> tuple[str, str, str]:
    cooperation = text(row.get("合作情况"))
    link = text(row.get("合作链接"))
    haystack = f"{cooperation} {notes_blob}".lower()
    if link:
        return "Published", "High", "合作链接存在"
    for stage, words in STAGE_RULES:
        for word in words:
            if word.lower() in haystack:
                confidence = "High" if cooperation and word.lower() in cooperation.lower() else "Medium"
                return stage, confidence, word
    return "", "Low", cooperation or notes_blob[:120]


def parse_quote(value: str) -> tuple[str, str, str]:
    if not value:
        return "", "", ""
    if "免费" in value:
        return "0", "CNY", ""
    match = re.search(r"(?P<currency>[$€£¥]|usd|eur|cny|rmb)?\s*(?P<amount>\d+(?:,\d{3})*(?:\.\d+)?)", value, re.I)
    if not match:
        return "", "", value
    currency = match.group("currency") or ""
    currency = {"$": "USD", "€": "EUR", "£": "GBP", "¥": "CNY", "rmb": "CNY"}.get(currency.lower(), currency.upper())
    return match.group("amount").replace(",", ""), currency, ""


def product_candidates(notes_blob: str) -> tuple[str, str, str]:
    candidates = []
    for match in PRODUCT_RE.findall(notes_blob):
        clean = re.sub(r"\s+", " ", match).strip()
        if clean and clean.lower() not in [x.lower() for x in candidates]:
            candidates.append(clean)
    if not candidates:
        return "", "Low", ""
    if len(candidates) == 1:
        return candidates[0], "Medium", candidates[0]
    return "; ".join(candidates), "Low", "多个产品线索"


def clean_notes(notes_blob: str, urls: list[str], product_text: str, stage_token: str) -> str:
    cleaned = notes_blob
    for url in urls:
        cleaned = cleaned.replace(url, "")
    if product_text:
        for product in product_text.split("; "):
            cleaned = re.sub(re.escape(product), "", cleaned, flags=re.I)
    if stage_token:
        cleaned = cleaned.replace(stage_token, "")
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ;；,，。")
    return cleaned


def build_cleaned_data() -> dict[str, Any]:
    headers, rows, raw_matrix = load_rows()
    media_map: OrderedDict[str, dict[str, Any]] = OrderedDict()
    media_row_counter = Counter()
    contacts = []
    campaigns = []
    review = []

    for index, row in enumerate(rows, start=1):
        media_key = normalize_media_key(row)
        media_row_counter[media_key] += 1
        media_id = f"M{list(media_map.keys()).index(media_key) + 1:04d}" if media_key in media_map else f"M{len(media_map) + 1:04d}"
        if media_key not in media_map:
            media_map[media_key] = {
                "media_id": media_id,
                "source_rows": str(row["raw_row"]),
                "name": text(row.get("名字")),
                "country": text(row.get("国家")),
                "parent_company": text(row.get("母公司")),
                "category": text(row.get("类目")),
                "platform_type": text(row.get("网站类型")),
                "website_url": text(row.get("链接")),
                "followers_or_traffic": text(row.get("流量/粉丝")),
                "dedupe_key_type": "name+link" if text(row.get("链接")) else "name+country",
                "dedupe_key": media_key,
                "media_review_note": "",
            }
        else:
            media_map[media_key]["source_rows"] += f", {row['raw_row']}"
            for target, source in [("country", "国家"), ("category", "类目"), ("platform_type", "网站类型"), ("followers_or_traffic", "流量/粉丝")]:
                if not media_map[media_key][target] and text(row.get(source)):
                    media_map[media_key][target] = text(row.get(source))

        contact_name, contact_role = split_contact(text(row.get("联系人&职位")))
        contact_info = text(row.get("联系方式"))
        emails = EMAIL_RE.findall(contact_info)
        phones = PHONE_RE.findall(contact_info)
        contacts.append(
            {
                "contact_id": f"C{len(contacts) + 1:04d}",
                "media_id": media_id,
                "source_row": row["raw_row"],
                "media_name": text(row.get("名字")),
                "contact_name": contact_name,
                "contact_role": contact_role,
                "email": emails[0] if emails else "",
                "phone": phones[0].strip() if phones else "",
                "telegram": contact_info if "telegram" in contact_info.lower() else "",
                "whatsapp": contact_info if "whatsapp" in contact_info.lower() else "",
                "brief_email": text(row.get("产品Brief 邮箱")),
                "press_release_email": text(row.get("Press release邮箱")),
                "contact_review_note": "" if contact_name or emails or phones else "联系人信息需确认",
                "raw_contact_info": contact_info,
            }
        )

        notes_blob = " ".join([text(row.get("合作情况")), text(row.get("合作备注")), text(row.get("合作备注2"))]).strip()
        all_urls = URL_RE.findall(notes_blob)
        stage, stage_confidence, stage_token = infer_stage(row, notes_blob)
        product_text, product_confidence, product_token = product_candidates(notes_blob)
        quote_amount, quote_currency, quote_note = parse_quote(text(row.get("报价")))
        deliverable = text(row.get("合作链接")) or next((u for u in all_urls if "youtube" in u.lower() or "youtu.be" in u.lower()), "")
        reference_urls = "; ".join([u for u in all_urls if u != deliverable])
        notes_clean = clean_notes(notes_blob, all_urls, product_text, stage_token)
        campaign = {
            "campaign_lead_id": f"L{len(campaigns) + 1:04d}",
            "media_id": media_id,
            "source_row": row["raw_row"],
            "media_name": text(row.get("名字")),
            "stage_candidate": stage,
            "stage_confidence": stage_confidence,
            "stage_review_note": "" if stage_confidence == "High" else stage_token,
            "possible_product_model": product_text,
            "product_confidence": product_confidence,
            "product_review_note": "" if product_confidence == "Medium" else product_token,
            "quotation_amount": quote_amount,
            "quotation_currency": quote_currency,
            "quotation_review_note": quote_note,
            "brief_sent": "Yes" if text(row.get("是否发产品Brief")) == "是" else ("No" if text(row.get("是否发产品Brief")) == "否" else ""),
            "deliverable_url_candidate": deliverable,
            "reference_urls": reference_urls,
            "clean_notes": notes_clean,
            "raw_cooperation": text(row.get("合作情况")),
            "raw_notes": text(row.get("合作备注")),
            "raw_notes2": text(row.get("合作备注2")),
        }
        campaigns.append(campaign)

        reasons = []
        if not text(row.get("名字")):
            reasons.append("缺少媒体名称")
        if stage_confidence != "High" and notes_blob:
            reasons.append("合作阶段需确认")
        if product_confidence == "Low" and product_text:
            reasons.append("产品线索多个或低置信度")
        if not contact_name and not emails and contact_info:
            reasons.append("联系人拆分需确认")
        if quote_note:
            reasons.append("报价需确认")
        if reasons:
            review.append(
                {
                    "review_id": f"R{len(review) + 1:04d}",
                    "source_row": row["raw_row"],
                    "media_id": media_id,
                    "media_name": text(row.get("名字")),
                    "review_reason": "; ".join(reasons),
                    "suggested_stage": stage,
                    "possible_product_model": product_text,
                    "original_text": notes_blob or contact_info,
                    "action": "",
                }
            )

    duplicate_groups = sum(1 for count in media_row_counter.values() if count > 1)
    media = list(media_map.values())
    summary = [
        ["Metric", "Value"],
        ["Raw rows processed", len(rows)],
        ["Media after dedupe", len(media)],
        ["Duplicate media groups", duplicate_groups],
        ["Contact rows", len(contacts)],
        ["Campaign lead rows", len(campaigns)],
        ["Needs Review rows", len(review)],
        ["Rows with possible product", sum(1 for c in campaigns if c["possible_product_model"])],
        ["Rows with URL in notes/link", sum(1 for c in campaigns if c["deliverable_url_candidate"] or c["reference_urls"])],
        ["High confidence stages", sum(1 for c in campaigns if c["stage_confidence"] == "High")],
        ["Medium confidence stages", sum(1 for c in campaigns if c["stage_confidence"] == "Medium")],
        ["Low confidence stages", sum(1 for c in campaigns if c["stage_confidence"] == "Low")],
    ]
    return {
        "source": str(SOURCE),
        "raw_headers": headers,
        "raw_matrix": raw_matrix,
        "media": media,
        "contacts": contacts,
        "campaigns": campaigns,
        "review": review,
        "summary": summary,
    }


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    data = build_cleaned_data()
    OUTPUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "output_json": str(OUTPUT_JSON),
        "raw_rows": len(data["raw_matrix"]) - 1,
        "media_rows": len(data["media"]),
        "contact_rows": len(data["contacts"]),
        "campaign_rows": len(data["campaigns"]),
        "review_rows": len(data["review"]),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
