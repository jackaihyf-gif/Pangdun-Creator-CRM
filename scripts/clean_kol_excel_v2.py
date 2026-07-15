from __future__ import annotations

import json
import re
from collections import Counter, OrderedDict, defaultdict
from pathlib import Path
from typing import Any

import openpyxl


SOURCE = Path(r"C:\Users\18620\Downloads\铭瑄红人记者库.xlsx")
OUTPUT_DIR = Path("outputs/cleaning")
OUTPUT_JSON = OUTPUT_DIR / "kol_clean_review_v2.json"

HEADERS = [
    "类目",
    "国家",
    "母公司",
    "名字",
    "流量/粉丝",
    "网站类型",
    "链接",
    "报价",
    "合作情况",
    "合作链接",
    "联系人&职位",
    "联系方式",
    "合作备注",
    "合作备注2",
    "是否发产品Brief",
    "产品Brief 邮箱",
    "Press release邮箱",
]

STAGE_OPTIONS = [
    "",
    "To Contact",
    "Contacted",
    "Waiting Reply",
    "Quoting",
    "Brief Sent",
    "Sample Sent",
    "In Production",
    "Published",
    "Paused",
    "Blacklisted",
]

STAGE_RULES = [
    ("Published", ["发布", "已发布"]),
    ("Brief Sent", ["已发brief", "已发产品brief", "发产品brief"]),
    ("In Production", ["测评", "评测", "制作中"]),
    ("Quoting", ["要钱", "报价", "收费", "付费"]),
    ("Waiting Reply", ["等回复", "待回复", "回复不活跃", "不活跃", "未回复"]),
    ("Contacted", ["3 logic联系", "已联系", "联系"]),
    ("To Contact", ["待开发", "开发"]),
]

EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+(?:\.[\w-]+)+")
PHONE_RE = re.compile(r"(?:\+?\d[\d\s().-]{6,}\d)")
URL_RE = re.compile(r"https?://[^\s)）]+", re.I)
PRODUCT_RE = re.compile(
    r"(?i)\b("
    r"DDR[45]"
    r"|B[0-9]{3}[A-Z0-9-]*(?:\s+(?:AIGA|WIFI|CROSS|PRO|MAX|ARCTIC|ICRAFT|TERMINATOR|CHALLENGER))*"
    r"|X[0-9]{3}[A-Z0-9-]*(?:\s+(?:AIGA|WIFI|CROSS|PRO|MAX|ARCTIC|ICRAFT|TERMINATOR|CHALLENGER))*"
    r"|Z[0-9]{3}[A-Z0-9-]*(?:\s+(?:AIGA|WIFI|CROSS|PRO|MAX|ARCTIC|ICRAFT|TERMINATOR|CHALLENGER))*"
    r"|H[0-9]{3}[A-Z0-9-]*(?:\s+(?:AIGA|WIFI|CROSS|PRO|MAX|ARCTIC|ICRAFT|TERMINATOR|CHALLENGER))*"
    r"|MS-(?:ICRAFT|TERMINATOR|GAMING|G-FORCE)[A-Z0-9 -]*"
    r"|(?:AIGA|ICRAFT|TERMINATOR|CHALLENGER)"
    r")\b"
)


def t(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_rows() -> tuple[list[list[Any]], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    raw_header = [t(v) for v in next(iterator)[: len(HEADERS)]]
    raw_matrix = [raw_header]
    rows: list[dict[str, Any]] = []
    carry = {"类目": "", "国家": "", "母公司": "", "名字": "", "链接": "", "网站类型": "", "流量/粉丝": ""}
    group_no = 0
    for row_number, values in enumerate(iterator, start=2):
        raw_values = list(values[: len(HEADERS)])
        if not any(v not in (None, "") for v in raw_values):
            continue
        raw_matrix.append(raw_values)
        row = {"raw_row": row_number}
        for index, header in enumerate(raw_header):
            row[header] = raw_values[index] if index < len(raw_values) else None
        if t(row.get("名字")):
            group_no += 1
            for key in carry:
                if t(row.get(key)):
                    carry[key] = t(row.get(key))
        row["group_no"] = group_no
        for key, value in carry.items():
            row[f"inherited_{key}"] = value if not t(row.get(key)) else t(row.get(key))
        rows.append(row)
    return raw_matrix, rows


def media_key(row: dict[str, Any]) -> str:
    name = row["inherited_名字"].lower()
    link = row["inherited_链接"].lower()
    country = row["inherited_国家"].lower()
    return f"{name}::{link or country}"


def compact_join(values: list[str], limit: int = 280) -> str:
    uniq = []
    for value in values:
        value = re.sub(r"\s+", " ", value).strip()
        if value and value not in uniq:
            uniq.append(value)
    joined = " | ".join(uniq)
    return joined[: limit - 1] + "…" if len(joined) > limit else joined


def split_contact(value: str) -> tuple[str, str]:
    value = re.sub(r"\s+", " ", value).strip()
    if not value:
        return "", ""
    for sep in ["｜", "|", " - ", "，", ",", "：", ":"]:
        if sep in value:
            left, right = value.split(sep, 1)
            return left.strip(), right.strip()
    words = value.split()
    role_words = {"editor", "writer", "journalist", "author", "reviewer", "本人", "编辑", "记者", "主编", "视频主"}
    if len(words) > 1 and any(w.lower() in role_words for w in words[-2:]):
        return " ".join(words[:-1]), words[-1]
    return value, ""


def infer_stage(rows: list[dict[str, Any]]) -> tuple[str, str, str]:
    source_text = " | ".join(t(row.get("合作情况")) for row in rows if t(row.get("合作情况")))
    links = [t(row.get("合作链接")) for row in rows if t(row.get("合作链接"))]
    if links:
        return "Published", "High", "有合作链接"
    lowered = source_text.lower()
    for stage, words in STAGE_RULES:
        matched = [word for word in words if word.lower() in lowered]
        if matched:
            return stage, "High", compact_join(matched)
    note_text = " | ".join(t(row.get("合作备注")) + " " + t(row.get("合作备注2")) for row in rows)
    lowered_notes = note_text.lower()
    for stage, words in STAGE_RULES:
        matched = [word for word in words if word.lower() in lowered_notes]
        if matched:
            return stage, "Needs Check", compact_join(matched)
    return "", "Needs Check", source_text[:160]


def product_suggestion(rows: list[dict[str, Any]]) -> tuple[str, str, str]:
    text_blob = " ".join(
        " ".join([t(row.get("合作情况")), t(row.get("合作备注")), t(row.get("合作备注2"))])
        for row in rows
    )
    text_blob = URL_RE.sub(" ", text_blob)
    found = []
    for match in PRODUCT_RE.findall(text_blob):
        value = re.sub(r"\s+", " ", match).strip()
        value = value.upper().replace("ICRAFT", "iCraft")
        if value and value not in found:
            found.append(value)
    if not found:
        return "", "Blank", ""
    if len(found) == 1:
        return found[0], "Suggest", "备注中只发现一个产品线索"
    return "; ".join(found), "Multiple - review", "备注中发现多个产品线索"


def stage_action(stage_confidence: str, product_confidence: str, rows: list[dict[str, Any]]) -> tuple[str, str]:
    reasons = []
    if stage_confidence != "High":
        reasons.append("阶段需确认")
    if product_confidence == "Multiple - review":
        reasons.append("产品多个需确认")
    if not rows[0]["inherited_名字"]:
        reasons.append("媒体名缺失")
    action = "确认后导入" if reasons else "可导入"
    return action, "；".join(reasons)


def extract_contacts(rows: list[dict[str, Any]], media_id: str, media_name: str) -> list[dict[str, Any]]:
    contacts = []
    seen = set()
    for row in rows:
        raw_contact = t(row.get("联系人&职位"))
        raw_info = t(row.get("联系方式"))
        if not raw_contact and not raw_info and not t(row.get("产品Brief 邮箱")) and not t(row.get("Press release邮箱")):
            continue
        name, role = split_contact(raw_contact)
        emails = EMAIL_RE.findall(raw_info)
        phones = PHONE_RE.findall(raw_info)
        key = (name.lower(), (emails[0].lower() if emails else ""), raw_info[:80])
        if key in seen:
            continue
        seen.add(key)
        contacts.append(
            {
                "media_id": media_id,
                "media_name": media_name,
                "source_row": row["raw_row"],
                "contact_name_final": name,
                "role_final": role,
                "email_final": emails[0] if emails else "",
                "phone_final": phones[0].strip() if phones else "",
                "telegram_final": raw_info if "telegram" in raw_info.lower() else "",
                "whatsapp_final": raw_info if "whatsapp" in raw_info.lower() else "",
                "brief_email_final": t(row.get("产品Brief 邮箱")),
                "press_release_email_final": t(row.get("Press release邮箱")),
                "contact_note": "" if (name or emails or phones) else "只保留原始联系方式，需人工拆分",
                "raw_contact_info": raw_info,
            }
        )
    return contacts


def build() -> dict[str, Any]:
    raw_matrix, rows = load_rows()
    groups: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for row in rows:
        key = media_key(row)
        groups.setdefault(key, []).append(row)

    review_rows = []
    contacts = []
    media_rows = []
    campaign_rows = []
    duplicate_groups = 0
    for index, (key, group_rows) in enumerate(groups.items(), start=1):
        if len(group_rows) > 1:
            duplicate_groups += 1
        media_id = f"M{index:04d}"
        first = group_rows[0]
        media_name = first["inherited_名字"]
        stage, stage_conf, stage_note = infer_stage(group_rows)
        product, product_conf, product_note = product_suggestion(group_rows)
        action, reasons = stage_action(stage_conf, product_conf, group_rows)
        urls = []
        deliverables = []
        for row in group_rows:
            deliverables.extend(URL_RE.findall(t(row.get("合作链接"))))
            urls.extend(URL_RE.findall(" ".join([t(row.get("合作情况")), t(row.get("合作备注")), t(row.get("合作备注2"))])))
        clean_note = compact_join([
            t(row.get("合作情况")),
            t(row.get("合作备注")),
            t(row.get("合作备注2")),
        ] for row in [])
        note_values = []
        for row in group_rows:
            note_values.extend([t(row.get("合作情况")), t(row.get("合作备注")), t(row.get("合作备注2"))])
        clean_note = compact_join(note_values, 420)
        media_rows.append(
            {
                "action": action,
                "review_reason": reasons,
                "media_id": media_id,
                "source_rows": compact_join([str(row["raw_row"]) for row in group_rows], 120),
                "media_name_final": media_name,
                "country_final": first["inherited_国家"],
                "category_final": first["inherited_类目"],
                "platform_type_final": first["inherited_网站类型"],
                "website_url_final": first["inherited_链接"],
                "followers_or_traffic_final": first["inherited_流量/粉丝"],
                "stage_final": stage,
                "stage_confidence": stage_conf,
                "product_final": product,
                "product_confidence": product_conf,
                "contacts_count": len(extract_contacts(group_rows, media_id, media_name)),
                "deliverable_url_candidate": compact_join(deliverables, 180),
                "reference_urls": compact_join([u for u in urls if u not in deliverables], 240),
                "clean_note_for_crm": clean_note,
                "raw_stage_note": stage_note,
            }
        )
        contact_rows = extract_contacts(group_rows, media_id, media_name)
        contacts.extend(contact_rows)
        campaign_rows.append(
            {
                "media_id": media_id,
                "media_name": media_name,
                "source_rows": compact_join([str(row["raw_row"]) for row in group_rows], 120),
                "stage_final": stage,
                "product_final": product,
                "brief_sent": "Yes" if any(t(row.get("是否发产品Brief")) == "是" for row in group_rows) else "",
                "quotation_raw": compact_join([t(row.get("报价")) for row in group_rows if t(row.get("报价"))], 160),
                "deliverable_url_candidate": compact_join(deliverables, 180),
                "campaign_note_final": clean_note,
            }
        )
        if action != "可导入":
            review_rows.append(
                {
                    "media_id": media_id,
                    "media_name": media_name,
                    "source_rows": compact_join([str(row["raw_row"]) for row in group_rows], 120),
                    "review_reason": reasons,
                    "suggested_stage": stage,
                    "suggested_product": product,
                    "what_to_do": "在人工确认主表中修正 action / stage_final / product_final / clean_note_for_crm",
                    "context_excerpt": clean_note,
                }
            )

    summary = [
        ["Metric", "Value"],
        ["Raw rows processed", len(rows)],
        ["Rows with blank media name inherited", sum(1 for row in rows if not t(row.get("名字")))],
        ["Media review rows", len(media_rows)],
        ["Duplicate/inherited media groups", duplicate_groups],
        ["Contact rows after contact-level dedupe", len(contacts)],
        ["Campaign rows", len(campaign_rows)],
        ["Review rows", len(review_rows)],
        ["Ready to import media rows", sum(1 for row in media_rows if row["action"] == "可导入")],
        ["Need confirmation media rows", sum(1 for row in media_rows if row["action"] != "可导入")],
        ["Rows with product suggestion", sum(1 for row in media_rows if row["product_final"])],
    ]
    return {
        "source": str(SOURCE),
        "raw_matrix": raw_matrix,
        "summary": summary,
        "media_review": media_rows,
        "contacts": contacts,
        "campaigns": campaign_rows,
        "needs_review": review_rows,
        "stage_options": STAGE_OPTIONS,
        "action_options": ["可导入", "确认后导入", "跳过", "合并到其他媒体"],
    }


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    data = build()
    OUTPUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "json": str(OUTPUT_JSON),
        "raw_rows": len(data["raw_matrix"]) - 1,
        "media_review_rows": len(data["media_review"]),
        "contacts": len(data["contacts"]),
        "needs_review": len(data["needs_review"]),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
