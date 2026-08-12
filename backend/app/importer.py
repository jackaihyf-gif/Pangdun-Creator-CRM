from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass
from io import BytesIO
from typing import Any

import openpyxl
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from .models import Campaign, Contact, Deliverable, ImportBatch, Media
from .media_taxonomy import normalize_media_payload


HEADERS = {
    "类目": "category",
    "国家": "country",
    "母公司": "parent_company",
    "名字": "name",
    "流量/粉丝": "followers_or_traffic",
    "网站类型": "platform_type",
    "链接": "website_url",
    "报价": "quotation",
    "合作情况": "cooperation",
    "合作链接": "deliverable_url",
    "联系人&职位": "contact_role",
    "联系方式": "contact_info",
    "合作备注": "notes",
    "合作备注2": "notes2",
    "是否发产品Brief": "brief_sent",
    "产品Brief 邮箱": "brief_email",
    "Press release邮箱": "press_release_email",
}
STAGE_MAP = {
    "待开发": "To Contact",
    "已联系": "Contacted",
    "等回复": "Waiting Reply",
    "报价": "Quoting",
    "要钱": "Quoting",
    "已发brief": "Brief Sent",
    "寄样": "Sample Sent",
    "制作": "In Production",
    "已产出": "Published",
    "拉黑": "Blacklisted",
    "暂停": "Paused",
}
EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+(?:\.[\w-]+)+")
PHONE_RE = re.compile(r"(?:\+?\d[\d\s().-]{6,}\d)")


@dataclass
class ImportResult:
    success_count: int
    skipped_count: int
    error_count: int
    errors: list[dict[str, Any]]
    rows: list[dict[str, Any]]
    created_count: int = 0
    updated_count: int = 0
    unchanged_count: int = 0
    conflict_count: int = 0
    batch_id: int | None = None


def load_rows(content: bytes) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    raw_rows = ws.iter_rows(values_only=True)
    headers = next(raw_rows, [])
    keys = [HEADERS.get(str(cell).strip(), None) if cell else None for cell in headers]
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(raw_rows, start=2):
        item: dict[str, Any] = {"row_number": row_number}
        for key, value in zip(keys, values):
            if key and value not in (None, ""):
                item[key] = str(value).strip() if not isinstance(value, (int, float)) else value
        if any(k for k in item.keys() if k != "row_number"):
            rows.append(normalize_row(item))
    return rows


def normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    contact_name, contact_role = split_contact_role(str(row.get("contact_role", "")).strip())
    contact_info = str(row.get("contact_info", "")).strip()
    emails = EMAIL_RE.findall(contact_info)
    phones = PHONE_RE.findall(contact_info)
    cooperation = str(row.get("cooperation", "")).strip()
    stage = infer_stage(cooperation)
    notes_parts = [row.get("notes"), row.get("notes2")]
    if cooperation and not stage:
        notes_parts.append(f"合作情况: {cooperation}")
    quotation = row.get("quotation")
    amount, currency, quote_note = parse_quotation(quotation)
    if quote_note:
        notes_parts.append(f"报价: {quote_note}")
    return {
        **row,
        "contact_name": contact_name,
        "contact_role_text": contact_role,
        "email": emails[0] if emails else None,
        "phone": phones[0].strip() if phones else None,
        "telegram": "Telegram" if "telegram" in contact_info.lower() else None,
        "contact_notes": contact_info if not emails and not phones else contact_info,
        "stage": stage or "Not Started",
        "quotation_amount": amount,
        "quotation_currency": currency,
        "campaign_notes": "\n".join(str(x) for x in notes_parts if x),
        "brief_sent_bool": str(row.get("brief_sent", "")).strip() == "是",
    }


def split_contact_role(value: str) -> tuple[str | None, str | None]:
    if not value:
        return None, None
    parts = value.split()
    if len(parts) <= 1:
        return value, None
    return " ".join(parts[:-1]), parts[-1]


def infer_stage(value: str) -> str | None:
    lowered = value.replace(" ", "").lower()
    for key, stage in STAGE_MAP.items():
        if key.lower() in lowered:
            return stage
    return None


def parse_quotation(value: Any) -> tuple[float | None, str | None, str | None]:
    if value in (None, ""):
        return None, None, None
    text = str(value).strip()
    if "免费" in text:
        return 0, "CNY", None
    match = re.search(r"(?P<currency>[$€£¥]|usd|eur|cny|rmb)?\s*(?P<amount>\d+(?:,\d{3})*(?:\.\d+)?)", text, re.I)
    if not match:
        return None, None, text
    currency = match.group("currency") or None
    currency = {"$": "USD", "€": "EUR", "£": "GBP", "¥": "CNY"}.get(currency, currency)
    return float(match.group("amount").replace(",", "")), currency.upper() if currency else None, None


def preview_import(content: bytes, db: Session | None = None) -> ImportResult:
    rows = load_rows(content)
    created = updated = unchanged = conflicts = 0
    previews = []
    for row in rows:
        action = "新增"
        if db:
            name = str(row.get("name", "")).strip()
            matches = find_media_matches(db, name, row.get("website_url"), row.get("country")) if name else []
            if len(matches) > 1:
                action = "冲突"
                conflicts += 1
                previews.append({**row, "import_action": action, "conflict_reason": f"匹配到 {len(matches)} 条媒体，需先合并或确认"})
                continue
            media = matches[0] if matches else None
            if media:
                contact = find_contact(db, media.id, row)
                action = "跳过" if contact else "更新"
                unchanged += int(action == "跳过")
                updated += int(action == "更新")
            else:
                created += 1
        else:
            created += 1
        previews.append({**row, "import_action": action})
    return ImportResult(len(rows), 0, 0, [], previews[:100], created, updated, unchanged, conflicts)


def confirm_import(db: Session, content: bytes, filename: str | None = None, user_id: int | None = None) -> ImportResult:
    rows = load_rows(content)
    success = skipped = errors = 0
    created = updated = unchanged = conflicts = 0
    error_rows: list[dict[str, Any]] = []
    preview_rows: list[dict[str, Any]] = []
    actions: list[dict[str, Any]] = []
    batch = ImportBatch(import_type="media", filename=filename, source_hash=hashlib.sha256(content).hexdigest(), user_id=user_id, status="processing")
    db.add(batch)
    db.flush()
    for row in rows:
        try:
            name = str(row.get("name", "")).strip()
            if not name:
                skipped += 1
                continue
            matches = find_media_matches(db, name, row.get("website_url"), row.get("country"))
            if len(matches) > 1:
                conflicts += 1
                skipped += 1
                error_rows.append({"row_number": row.get("row_number"), "error": f"匹配到 {len(matches)} 条媒体，本行未导入"})
                continue
            media = matches[0] if matches else None
            if not media:
                media_data = normalize_media_payload(dict(
                    name=name,
                    country=row.get("country"),
                    category=row.get("category"),
                    platform_type=row.get("platform_type"),
                    website_url=row.get("website_url"),
                    followers_or_traffic=safe_k(row.get("followers_or_traffic")),
                    cooperation_status=row.get("cooperation"),
                    notes=row.get("parent_company"),
                ))
                media = Media(**media_data)
                db.add(media)
                db.flush()
                actions.append({"kind": "create", "entity": "media", "id": media.id})
                created += 1
            else:
                media_changes = fill_missing(media, normalize_media_payload({
                    "name": name, "country": row.get("country"), "category": row.get("category"),
                    "platform_type": row.get("platform_type"), "website_url": row.get("website_url"),
                    "followers_or_traffic": safe_k(row.get("followers_or_traffic")),
                    "cooperation_status": row.get("cooperation"), "notes": row.get("parent_company"),
                }), ["country", "category", "platform_type", "website_url", "followers_or_traffic", "media_tier", "cooperation_status", "notes"])
                if media_changes:
                    actions.append({"kind": "update", "entity": "media", "id": media.id, "before": media_changes})
                    updated += 1
            contact_data = {
                "name": row.get("contact_name"), "role": row.get("contact_role_text"), "email": row.get("email"),
                "phone": row.get("phone"), "telegram": row.get("telegram"), "brief_email": row.get("brief_email"),
                "press_release_email": row.get("press_release_email"), "notes": row.get("contact_notes"),
            }
            contact = find_contact(db, media.id, row)
            if contact:
                contact_changes = fill_missing(contact, contact_data, list(contact_data))
                if contact_changes:
                    actions.append({"kind": "update", "entity": "contact", "id": contact.id, "before": contact_changes})
                    updated += 1
                else:
                    unchanged += 1
            elif any(contact_data.values()):
                contact = Contact(media_id=media.id, **contact_data)
                db.add(contact)
                db.flush()
                actions.append({"kind": "create", "entity": "contact", "id": contact.id})
                created += 1
            campaign = None
            if row.get("campaign_notes") or row.get("quotation_amount") is not None or row.get("brief_sent_bool") or row.get("deliverable_url"):
                campaign = db.query(Campaign).filter(
                    Campaign.media_id == media.id,
                    Campaign.project_id.is_(None),
                    Campaign.stage == row.get("stage", "Not Started"),
                    Campaign.notes == row.get("campaign_notes"),
                    Campaign.quotation_amount == row.get("quotation_amount"),
                ).first()
                if not campaign:
                    campaign = Campaign(media_id=media.id, stage=row.get("stage", "Not Started"), quotation_amount=row.get("quotation_amount"), quotation_currency=row.get("quotation_currency"), brief_sent=row.get("brief_sent_bool", False), notes=row.get("campaign_notes"))
                    db.add(campaign)
                    db.flush()
                    actions.append({"kind": "create", "entity": "campaign", "id": campaign.id})
                    created += 1
                else:
                    unchanged += 1
            if campaign and row.get("deliverable_url"):
                deliverable = db.query(Deliverable).filter(Deliverable.campaign_id == campaign.id, Deliverable.url == row.get("deliverable_url")).first()
                if not deliverable:
                    deliverable = Deliverable(campaign_id=campaign.id, url=row.get("deliverable_url"), deliverable_type="Other")
                    db.add(deliverable)
                    db.flush()
                    actions.append({"kind": "create", "entity": "deliverable", "id": deliverable.id})
                    created += 1
                else:
                    unchanged += 1
            success += 1
            if len(preview_rows) < 100:
                preview_rows.append(row)
        except Exception as exc:
            errors += 1
            error_rows.append({"row_number": row.get("row_number"), "error": str(exc)})
    summary = {"success_count": success, "skipped_count": skipped, "error_count": errors, "created_count": created, "updated_count": updated, "unchanged_count": unchanged, "conflict_count": conflicts}
    batch.status = "completed" if not errors else "completed_with_errors"
    batch.summary_json = json.dumps(summary, ensure_ascii=False)
    batch.undo_json = json.dumps(actions, ensure_ascii=False)
    db.commit()
    return ImportResult(success, skipped, errors, error_rows, preview_rows, created, updated, unchanged, conflicts, batch.id)


def normalized_identity(value: str | None) -> str:
    return re.sub(r"[^\w@]+", "", (value or "").strip().lower(), flags=re.UNICODE)


def find_contact(db: Session, media_id: int, row: dict[str, Any]) -> Contact | None:
    email = normalized_identity(row.get("email"))
    phone = normalized_identity(row.get("phone"))
    name = normalized_identity(row.get("contact_name"))
    candidates = db.query(Contact).filter(Contact.media_id == media_id).all()
    if email:
        match = next((item for item in candidates if normalized_identity(item.email) == email), None)
        if match:
            return match
    if phone:
        match = next((item for item in candidates if normalized_identity(item.phone) == phone), None)
        if match:
            return match
    return next((item for item in candidates if name and normalized_identity(item.name) == name), None)


def fill_missing(item: Any, data: dict[str, Any], fields: list[str]) -> dict[str, Any]:
    before: dict[str, Any] = {}
    for field in fields:
        current = getattr(item, field, None)
        incoming = data.get(field)
        if current in (None, "") and incoming not in (None, ""):
            before[field] = current
            setattr(item, field, incoming)
    return before


def find_media(db: Session, name: str, website_url: str | None, country: str | None) -> Media | None:
    matches = find_media_matches(db, name, website_url, country)
    return matches[0] if matches else None


def find_media_matches(db: Session, name: str, website_url: str | None, country: str | None) -> list[Media]:
    if website_url:
        return db.query(Media).filter(func.lower(Media.name) == name.lower(), Media.website_url == website_url).all()
    return db.query(Media).filter(func.lower(Media.name) == name.lower(), or_(Media.country == country, Media.country.is_(None))).all()


def safe_int(value: Any) -> int | None:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def safe_k(value: Any) -> float | None:
    try:
        return round(float(value) / 1000, 2)
    except (TypeError, ValueError):
        return None
