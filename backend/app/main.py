from __future__ import annotations

import json
import re
from contextlib import asynccontextmanager
from difflib import SequenceMatcher
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Annotated, Any
from urllib.parse import quote, unquote

import httpx
from fastapi import Depends, FastAPI, File, Header, HTTPException, Query, Response, UploadFile
from fastapi.encoders import jsonable_encoder
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from fastapi.staticfiles import StaticFiles
from sqlalchemy import String, and_, func, or_
from sqlalchemy.orm import Session, joinedload

from .agent_service import AgentConfigurationError, AgentSourceError, agent_config, deepseek_json_extract, fetch_public_source, source_hash
from .auth import CLI_TOKEN_EXPIRE_DAYS, clear_session_cookie, create_cli_token, current_user, hash_password, require_roles, set_session_cookie, verify_password
from .database import Base, apply_compat_migrations, engine, get_db
from .content_monitor_service import monitor_runtime_status, run_content_monitor, start_content_monitor, stop_content_monitor
from .execution_importer import confirm_execution_import, preview_execution_import
from .importer import confirm_import, preview_import
from .models import Activity, AgentRun, AuditLog, Campaign, CampaignStageEvent, Contact, CostItem, Deliverable, DeliverablePerformanceSnapshot, ImportBatch, Media, Product, Project, ProjectProduct, Shipment, ShipmentItem, ShippingAddress, User
from .media_taxonomy import COOPERATION_STATUSES, COUNTRIES, MEDIA_CHANNELS, VERIFICATION_STATUSES, infer_audience_metric_type, metric_value_in_k, normalize_channel, normalize_cooperation_status, normalize_country, normalize_media_payload
from .profile_links import clean_profile_links, profile_identity
from .social_identity_service import SocialIdentityError, fetch_social_identity, merge_social_identity_proposal, social_platform
from .youtube_service import YouTubeConfigurationError, YouTubeSourceError, fetch_youtube_channel, is_youtube_url, merge_youtube_proposal
from .product_backfill import backfill_products, ensure_project_link, find_or_create_product, find_product_matches, product_aliases, product_identity
from .schemas import (
    CampaignBase,
    CampaignOut,
    ContactBase,
    ContactOut,
    DeliverableBase,
    DeliverableOut,
    LoginIn,
    MediaBase,
    MediaReviewResolveIn,
    MediaReviewBatchIn,
    AgentExtractIn,
    AgentApplyIn,
    AgentRejectIn,
    MediaMergeIn,
    MediaOut,
    ProductBase,
    ProductMergeIn,
    ProductOut,
    ProjectBase,
    ProjectOut,
    ShipmentBase,
    ShipmentOut,
    ShippingAddressBase,
    ShippingAddressOut,
    CostItemBase,
    CostItemOut,
    ActivityBase,
    ActivityOut,
    CollaborationBulkPatch,
    CollaborationAdvanceIn,
    CollaborationStatusActionIn,
    CollaborationPatch,
    ProjectShipmentBase,
    UserCreate,
    UserOut,
    UserUpdate,
)


ROOT = Path(__file__).resolve().parents[2]
FRONTEND_DIST = ROOT / "outputs" / "frontend-dist-v2"
if not (FRONTEND_DIST / "assets").exists():
    FRONTEND_DIST = ROOT / "frontend" / "dist"
STAGES = {
    "Not Started",
    "To Contact",
    "Contacted",
    "Waiting Reply",
    "Quoting",
    "Brief Sent",
    "Sample Sent",
    "In Production",
    "Published",
    "Closed",
    "Paused",
    "Blacklisted",
}
SAMPLE_STATUSES = {
    "Not Needed",
    "Not Sent",
    "Preparing",
    "Shipped",
    "In Transit",
    "Customs Clearance",
    "Delivered",
    "Issue",
}
DELIVERABLE_TYPES = {
    "YouTube Video",
    "YouTube Shorts",
    "Website Article",
    "Instagram Reel",
    "TikTok Video",
    "Press Release",
    "Other",
}
EXECUTION_STATUSES = {"待确认", "待发货", "运输中", "已签收待产出", "内容审核中", "已发布", "已结算", "已暂停", "已取消", "已暂停/取消"}
EXECUTION_SEQUENCE = ["待确认", "待发货", "运输中", "已签收待产出", "内容审核中", "已发布", "已结算"]
NEXT_ACTION_BY_STATUS = {
    "待确认": "确认合作意向与报价",
    "待发货": "确认收件信息并安排寄样",
    "运输中": "跟踪物流并同步预计到达时间",
    "已签收待产出": "确认内容排期与脚本方向",
    "内容审核中": "完成内容审核并反馈修改意见",
    "已发布": "回收内容链接与效果数据",
    "已结算": "归档合作结果与复盘",
    "已暂停": "确认恢复条件与下一次检查时间",
    "已取消": "归档取消原因与合作结论",
    "已暂停/取消": "确认恢复条件与下一次检查时间",
}
FOLLOW_UP_CLOSED_STATUSES = {"已结算", "已暂停", "已取消", "已暂停/取消"}

SHIPMENT_STATUS_PROGRESS = {
    "待发货": 1,
    "运输中": 2,
    "已签收待产出": 3,
}
HISTORICAL_PROJECT_PREFIX = "历史导入"


def collaboration_workflow_health(item: Campaign) -> dict[str, Any]:
    if item.execution_status in FOLLOW_UP_CLOSED_STATUSES:
        return {"workflow_health": "closed", "workflow_label": "无需跟进", "workflow_warnings": []}
    if item.follow_up_done:
        return {
            "workflow_health": "needs_next_step",
            "workflow_label": "待续排",
            "workflow_warnings": ["当前待办已完成，请安排新的下一步行动和日期"],
        }
    missing_action = not (item.next_action or "").strip()
    missing_date = item.follow_up_date is None
    if missing_action and missing_date:
        return {
            "workflow_health": "missing_both",
            "workflow_label": "待补行动/日期",
            "workflow_warnings": ["缺少下一步行动", "缺少跟进日期"],
        }
    if missing_action:
        return {"workflow_health": "missing_action", "workflow_label": "待补行动", "workflow_warnings": ["缺少下一步行动"]}
    if missing_date:
        return {"workflow_health": "missing_date", "workflow_label": "待排期", "workflow_warnings": ["缺少跟进日期"]}
    if item.follow_up_date < date.today():
        return {"workflow_health": "overdue", "workflow_label": "已逾期", "workflow_warnings": ["跟进日期已逾期"]}
    return {"workflow_health": "ready", "workflow_label": "已安排", "workflow_warnings": []}


def collaboration_advance_state(item: Campaign) -> dict[str, Any]:
    if item.execution_status not in EXECUTION_SEQUENCE:
        return {"next_status": None, "advance_ready": False, "advance_blockers": [], "advance_requirements": []}
    current_index = EXECUTION_SEQUENCE.index(item.execution_status)
    if current_index == len(EXECUTION_SEQUENCE) - 1:
        return {"next_status": None, "advance_ready": False, "advance_blockers": [], "advance_requirements": []}
    target = EXECUTION_SEQUENCE[current_index + 1]
    blockers: list[str] = []
    requirements: list[str] = []
    shipments = item.shipments or []
    deliverables = item.deliverables or []
    costs = item.cost_items or []
    if target == "运输中" and not any((row.tracking_number or "").strip() for row in shipments):
        blockers.append("缺少物流单号")
        requirements.append("tracking_number")
    elif target == "已签收待产出" and not any(row.delivered_at for row in shipments):
        blockers.append("尚未确认签收日期")
        requirements.append("delivered_at")
    elif target == "内容审核中" and not deliverables:
        blockers.append("尚未登记待审核内容")
        requirements.append("content_title")
    elif target == "已发布":
        if not any((row.url or "").strip() for row in deliverables):
            blockers.append("缺少发布链接")
            requirements.append("publication_url")
        if not any(row.published_at for row in deliverables):
            blockers.append("缺少发布日期")
            requirements.append("published_at")
    elif target == "已结算":
        if not costs:
            blockers.append("尚未登记费用；如本次无需付款，请明确确认")
            requirements.append("no_payment_required")
        elif any(row.payment_status not in {"已付款", "无需付款"} for row in costs):
            blockers.append("仍有未完成付款的费用记录")
            requirements.append("settle_costs")
    return {
        "next_status": target,
        "advance_ready": not blockers,
        "advance_blockers": blockers,
        "advance_requirements": requirements,
    }


def collaboration_stage_metadata(item: Campaign) -> dict[str, Any]:
    entered_at = item.execution_status_changed_at or item.updated_at or item.created_at
    elapsed = max(0, (datetime.utcnow() - entered_at).days) if entered_at else 0
    return {"status_entered_at": entered_at, "days_in_status": elapsed}


def set_campaign_status(
    db: Session,
    item: Campaign,
    user: User | None,
    target_status: str,
    action: str,
    reason: str | None = None,
) -> None:
    before_status = item.execution_status
    if target_status == before_status:
        return
    item.execution_status = target_status
    item.execution_status_changed_at = datetime.utcnow()
    item.next_action = NEXT_ACTION_BY_STATUS.get(target_status)
    item.follow_up_done = target_status == "已结算"
    if target_status == "已发布":
        item.stage = "Published"
    elif target_status == "已结算":
        item.stage = "Closed"
    elif target_status in {"已暂停", "已暂停/取消"}:
        item.stage = "Paused"
    elif target_status == "已取消":
        item.stage = "Closed"
    elif item.stage in {"Published", "Closed", "Paused"}:
        item.stage = "Not Started"
    db.add(CampaignStageEvent(campaign_id=item.id, user_id=user.id if user else None, from_status=before_status, to_status=target_status, action=action, reason=reason))
    label = {"advance": "阶段推进", "pause": "暂停执行", "cancel": "取消合作", "resume": "恢复执行", "rollback": "阶段回退", "undo": "撤销推进", "override": "管理员调整", "shipment_sync": "物流同步"}.get(action, "状态更新")
    detail = f"执行阶段由“{before_status}”变更为“{target_status}”"
    if reason:
        detail += f"；原因：{reason}"
    db.add(Activity(campaign_id=item.id, user_id=user.id if user else None, activity_type=label, content=detail))

Base.metadata.create_all(bind=engine)
apply_compat_migrations()
with next(get_db()) as bootstrap_db:
    backfill_products(bootstrap_db)

@asynccontextmanager
async def app_lifespan(_: FastAPI):
    start_content_monitor()
    try:
        yield
    finally:
        stop_content_monitor()


app = FastAPI(title="Pangdun KOL CRM", lifespan=app_lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://127.0.0.1:5173", "http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def list_payload(query, page: int, page_size: int):
    total = query.count()
    items = query.offset((page - 1) * page_size).limit(page_size).all()
    return {"items": jsonable_encoder(items), "total": total}


def editable_user(user: Annotated[User, Depends(require_roles("Admin", "Editor"))]) -> User:
    return user


def add_audit_log(
    db: Session,
    user: User,
    action: str,
    entity_type: str,
    entity_id: int | str | None,
    before: Any = None,
    after: Any = None,
    reason: str | None = None,
) -> None:
    db.add(AuditLog(
        user_id=user.id,
        action=action,
        entity_type=entity_type,
        entity_id=str(entity_id) if entity_id is not None else None,
        before_json=json.dumps(jsonable_encoder(before), ensure_ascii=False) if before is not None else None,
        after_json=json.dumps(jsonable_encoder(after), ensure_ascii=False) if after is not None else None,
        reason=unquote((reason or "").strip()) or None,
    ))


def validate_campaign(payload: CampaignBase) -> None:
    if payload.stage not in STAGES:
        raise HTTPException(400, f"Invalid stage: {payload.stage}")
    if payload.sample_status not in SAMPLE_STATUSES:
        raise HTTPException(400, f"Invalid sample status: {payload.sample_status}")
    if payload.execution_status not in EXECUTION_STATUSES:
        raise HTTPException(400, f"Invalid execution status: {payload.execution_status}")


@app.get("/api/health")
def health():
    return {"ok": True}


@app.post("/api/auth/login", response_model=UserOut)
def login(payload: LoginIn, response: Response, db: Annotated[Session, Depends(get_db)]):
    user = db.query(User).filter(func.lower(User.email) == payload.email.lower()).first()
    if not user or not user.is_active or not verify_password(payload.password, user.password_hash):
        raise HTTPException(401, "Invalid email or password")
    set_session_cookie(response, user)
    return user


@app.post("/api/auth/cli-token")
def cli_token(payload: LoginIn, db: Annotated[Session, Depends(get_db)]):
    user = db.query(User).filter(func.lower(User.email) == payload.email.lower()).first()
    if not user or not user.is_active or not verify_password(payload.password, user.password_hash):
        raise HTTPException(401, "Invalid email or password")
    return {
        "access_token": create_cli_token(user),
        "token_type": "bearer",
        "expires_in_days": CLI_TOKEN_EXPIRE_DAYS,
        "user": UserOut.model_validate(user).model_dump(mode="json"),
    }


@app.post("/api/auth/logout")
def logout(response: Response):
    clear_session_cookie(response)
    return {"ok": True}


@app.get("/api/auth/me", response_model=UserOut)
def me(user: Annotated[User, Depends(current_user)]):
    return user


@app.get("/api/options")
def options(user: Annotated[User, Depends(current_user)]):
    return {
        "roles": ["Admin", "Editor", "Viewer"],
        "stages": sorted(STAGES),
        "sample_statuses": sorted(SAMPLE_STATUSES),
        "deliverable_types": sorted(DELIVERABLE_TYPES),
        "execution_statuses": sorted(EXECUTION_STATUSES),
        "payment_statuses": ["未付款", "部分付款", "已付款", "无需付款"],
        "media_channels": MEDIA_CHANNELS,
        "cooperation_statuses": COOPERATION_STATUSES,
        "verification_statuses": VERIFICATION_STATUSES,
        "countries": [{"code": item["code"], "label": item["label"]} for item in COUNTRIES],
    }


@app.get("/api/users", response_model=dict)
def list_users(db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))]):
    items = db.query(User).order_by(User.created_at.desc()).all()
    return {"items": jsonable_encoder(items), "total": db.query(User).count()}


@app.post("/api/users", response_model=UserOut)
def create_user(payload: UserCreate, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))]):
    if db.query(User).filter(func.lower(User.email) == payload.email.lower()).first():
        raise HTTPException(400, "Email already exists")
    item = User(email=payload.email.lower(), name=payload.name, role=payload.role, password_hash=hash_password(payload.password))
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@app.put("/api/users/{item_id}", response_model=UserOut)
def update_user(item_id: int, payload: UserUpdate, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))]):
    item = db.get(User, item_id)
    if not item:
        raise HTTPException(404, "User not found")
    data = payload.model_dump(exclude_unset=True)
    if "password" in data and data["password"]:
        item.password_hash = hash_password(data.pop("password"))
    for key, value in data.items():
        setattr(item, key, value)
    db.commit()
    db.refresh(item)
    return item


@app.get("/api/audit-logs")
def list_audit_logs(
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(require_roles("Admin", "Editor"))],
    limit: int = Query(default=50, ge=1, le=500),
    entity_type: str | None = None,
    entity_id: str | None = None,
):
    query = db.query(AuditLog).options(joinedload(AuditLog.user))
    if entity_type:
        query = query.filter(AuditLog.entity_type == entity_type)
    if entity_id:
        query = query.filter(AuditLog.entity_id == entity_id)
    rows = query.order_by(AuditLog.created_at.desc()).limit(limit).all()
    return {"items": [{
        "id": row.id,
        "user": row.user.name if row.user else None,
        "action": row.action,
        "entity_type": row.entity_type,
        "entity_id": row.entity_id,
        "before": json.loads(row.before_json) if row.before_json else None,
        "after": json.loads(row.after_json) if row.after_json else None,
        "reason": row.reason,
        "created_at": row.created_at,
    } for row in rows]}


@app.post("/api/audit-logs/{log_id}/restore")
def restore_audit_log(
    log_id: int,
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(require_roles("Admin"))],
):
    log = db.get(AuditLog, log_id)
    if not log or log.entity_type != "media" or log.action not in {"update", "restore"} or not log.before_json:
        raise HTTPException(400, "该记录不支持单条恢复")
    item = db.get(Media, int(log.entity_id or 0))
    if not item:
        raise HTTPException(404, "Media not found")
    current = MediaOut.model_validate(item).model_dump(mode="json")
    restored = MediaBase.model_validate(json.loads(log.before_json)).model_dump()
    restored["profile_links"] = clean_profile_links(restored.get("profile_links"), restored.get("website_url"))
    require_unique_media_identity(db, restored["profile_links"], item.id)
    restored["website_url"] = restored["profile_links"][0]["url"] if restored["profile_links"] else None
    for key, value in restored.items():
        setattr(item, key, value)
    add_audit_log(db, user, "restore", "media", item.id, before=current, after=restored, reason=f"恢复审计记录 #{log.id}")
    db.commit()
    db.refresh(item)
    return MediaOut.model_validate(item)


def agent_run_payload(row: AgentRun) -> dict[str, Any]:
    return {
        "id": row.id,
        "task_type": row.task_type,
        "input_type": row.input_type,
        "source_label": row.source_label,
        "status": row.status,
        "model": row.model,
        "proposal": json.loads(row.proposal_json) if row.proposal_json else None,
        "usage": json.loads(row.usage_json) if row.usage_json else None,
        "error_message": row.error_message,
        "target_media_id": row.target_media_id,
        "user": row.user.name if row.user else None,
        "reviewed_by": row.reviewed_by.name if row.reviewed_by else None,
        "created_at": row.created_at,
        "reviewed_at": row.reviewed_at,
    }


@app.get("/api/agent/status")
def agent_status(user: Annotated[User, Depends(current_user)]):
    return agent_config()


@app.get("/api/agent/runs")
def list_agent_runs(
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(current_user)],
    status: str | None = None,
    limit: int = Query(default=30, ge=1, le=100),
):
    query = db.query(AgentRun).options(joinedload(AgentRun.user), joinedload(AgentRun.reviewed_by))
    if status:
        query = query.filter(AgentRun.status == status)
    rows = query.order_by(AgentRun.created_at.desc()).limit(limit).all()
    return {"items": [agent_run_payload(row) for row in rows], "total": query.count()}


@app.post("/api/agent/extract")
def extract_with_agent(
    payload: AgentExtractIn,
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(require_roles("Admin", "Editor"))],
):
    if payload.input_type not in {"text", "url"}:
        raise HTTPException(400, "Agent 首版仅支持网页 URL 或粘贴文本")
    raw = payload.content.strip()
    if len(raw) < 10:
        raise HTTPException(400, "请提供足够的网页、Media Kit、邮件或表格文字")
    source_label = (payload.source_label or (raw if payload.input_type == "url" else "粘贴文本")).strip()
    try:
        if payload.input_type == "url":
            if is_youtube_url(raw):
                youtube_channel = fetch_youtube_channel(raw)
                social_identity = None
                content = youtube_channel.source_text()
                source_label = payload.source_label or youtube_channel.canonical_url
            elif social_platform(raw):
                youtube_channel = None
                social_identity = fetch_social_identity(raw)
                content = social_identity.source_text()
                source_label = payload.source_label or social_identity.canonical_url
            else:
                youtube_channel = None
                social_identity = None
                content, final_url = fetch_public_source(raw)
                source_label = payload.source_label or final_url
        else:
            youtube_channel = None
            social_identity = None
            content = raw[:45_000]
    except YouTubeConfigurationError as exc:
        raise HTTPException(503, str(exc)) from exc
    except (AgentSourceError, SocialIdentityError, YouTubeSourceError, httpx.HTTPError) as exc:
        raise HTTPException(400, f"读取来源失败：{exc}") from exc
    run = AgentRun(
        input_type=payload.input_type,
        source_label=source_label[:500],
        source_hash=source_hash(content),
        status="processing",
        model=agent_config()["model"],
        user_id=user.id,
    )
    db.add(run)
    db.commit()
    db.refresh(run)
    try:
        proposal, usage = deepseek_json_extract(content, source_label)
        if youtube_channel:
            proposal = merge_youtube_proposal(proposal, youtube_channel)
        elif social_identity:
            proposal = merge_social_identity_proposal(proposal, social_identity)
        proposed_media = proposal.get("media") or {}
        links = proposed_media.get("profile_links") or []
        matches = media_identity_matches(db, links)
        if not matches and proposed_media.get("name"):
            normalized_name = proposed_media["name"].strip().casefold()
            name_rows = db.query(Media).filter(func.lower(Media.name) == normalized_name).limit(10).all()
            matches = [{"id": item.id, "name": item.name, "reason": "媒体名称相同"} for item in name_rows]
        proposal["match_candidates"] = matches
        proposal["suggested_target_media_id"] = matches[0]["id"] if len(matches) == 1 else None
        run.target_media_id = proposal["suggested_target_media_id"]
        run.proposal_json = json.dumps(jsonable_encoder(proposal), ensure_ascii=False)
        run.usage_json = json.dumps(usage, ensure_ascii=False)
        run.status = "proposed"
        db.commit()
        db.refresh(run)
        return agent_run_payload(run)
    except AgentConfigurationError as exc:
        run.status = "failed"
        run.error_message = str(exc)
        db.commit()
        raise HTTPException(503, str(exc)) from exc
    except (httpx.HTTPError, RuntimeError, ValueError) as exc:
        run.status = "failed"
        run.error_message = str(exc)[:1000]
        db.commit()
        raise HTTPException(502, f"Agent 提取失败：{exc}") from exc


AGENT_MEDIA_FIELDS = {
    "name", "country", "platform_type", "category", "profile_links",
    "followers_or_traffic", "audience_metric_type", "metric_source", "metric_verified_at",
    "cooperation_status", "notes",
}
AGENT_CONTACT_FIELDS = {"name", "role", "email", "phone", "whatsapp", "telegram"}


@app.post("/api/agent/runs/{run_id}/apply")
def apply_agent_run(
    run_id: int,
    payload: AgentApplyIn,
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(require_roles("Admin", "Editor"))],
):
    run = db.get(AgentRun, run_id)
    if not run or run.status != "proposed" or not run.proposal_json:
        raise HTTPException(400, "该 Agent 建议不存在或已经处理")
    proposal = json.loads(run.proposal_json)
    selected = set(payload.selected_fields)
    if not selected:
        raise HTTPException(400, "请至少选择一个要写入的字段")
    proposed_media = proposal.get("media") or {}
    valid_selected = {
        f"media.{field}" for field in AGENT_MEDIA_FIELDS
        if proposed_media.get(field) not in (None, "", [])
    }
    for index, proposed_contact in enumerate(proposal.get("contacts") or []):
        valid_selected.update(
            f"contacts.{index}.{field}" for field in AGENT_CONTACT_FIELDS
            if proposed_contact.get(field) not in (None, "")
        )
    if not selected.issubset(valid_selected):
        raise HTTPException(400, "选择中包含 Agent 未建议或不允许写入的字段")
    target_id = payload.target_media_id or run.target_media_id
    item = db.get(Media, target_id) if target_id else None
    if target_id and not item:
        raise HTTPException(404, "Target media not found")
    if not item and not payload.create_media:
        raise HTTPException(400, "请选择现有媒体，或确认创建新媒体")
    media_changes = {
        field: proposed_media.get(field)
        for field in AGENT_MEDIA_FIELDS
        if f"media.{field}" in selected and proposed_media.get(field) not in (None, "", [])
    }
    confidence = proposal.get("confidence") or {}
    selected_scores = [float(confidence[key]) for key in selected if key in confidence]
    overall_confidence = round(sum(selected_scores) / len(selected_scores), 3) if selected_scores else 0.49
    source_label = run.source_label or "Agent 提取"
    try:
        if item:
            before = MediaOut.model_validate(item).model_dump(mode="json")
            merged = {**before, **media_changes}
        else:
            if not media_changes.get("name"):
                raise HTTPException(400, "新建媒体必须选择名称字段")
            before = None
            merged = media_changes
        merged.update({
            "data_source": source_label,
            "data_capture_method": "agent",
            "data_confidence": overall_confidence,
            "last_verified_at": date.today() if overall_confidence >= 0.8 else None,
            "verification_status": "部分核验" if overall_confidence >= 0.8 else "待核验",
        })
        normalized = normalize_media_payload(MediaBase.model_validate(merged).model_dump())
        normalized["profile_links"] = clean_profile_links(normalized.get("profile_links"), normalized.get("website_url"))
        require_unique_media_identity(db, normalized["profile_links"], item.id if item else None)
        normalized["website_url"] = normalized["profile_links"][0]["url"] if normalized["profile_links"] else None
        if item:
            for key, value in normalized.items():
                setattr(item, key, value)
        else:
            item = Media(**normalized)
            db.add(item)
            db.flush()
        created_contacts = 0
        for index, proposed_contact in enumerate(proposal.get("contacts") or []):
            contact_data = {
                field: proposed_contact.get(field)
                for field in AGENT_CONTACT_FIELDS
                if f"contacts.{index}.{field}" in selected and proposed_contact.get(field) not in (None, "")
            }
            if not contact_data:
                continue
            contact_data.update({
                "media_id": item.id,
                "data_source": source_label,
                "data_capture_method": "agent",
                "data_confidence": overall_confidence,
                "verified_at": date.today() if overall_confidence >= 0.8 else None,
            })
            contact_data = normalize_contact_data(contact_data)
            require_unique_contact(db, contact_data)
            db.add(Contact(**contact_data))
            created_contacts += 1
        after = MediaOut.model_validate(item).model_dump(mode="json")
        add_audit_log(db, user, "agent_apply", "media", item.id, before=before, after={"selected_fields": sorted(selected), "media": after, "contacts_created": created_contacts, "agent_run_id": run.id}, reason=f"人工确认 Agent 建议；来源={source_label}")
        run.status = "applied"
        run.target_media_id = item.id
        run.reviewed_by_id = user.id
        run.reviewed_at = datetime.utcnow()
        db.commit()
        db.refresh(item)
        return {"ok": True, "media": MediaOut.model_validate(item), "contacts_created": created_contacts, "run": agent_run_payload(run)}
    except Exception:
        db.rollback()
        raise


@app.post("/api/agent/runs/{run_id}/reject")
def reject_agent_run(
    run_id: int,
    payload: AgentRejectIn,
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(require_roles("Admin", "Editor"))],
):
    run = db.get(AgentRun, run_id)
    if not run or run.status != "proposed":
        raise HTTPException(400, "该 Agent 建议不存在或已经处理")
    run.status = "rejected"
    run.reviewed_by_id = user.id
    run.reviewed_at = datetime.utcnow()
    add_audit_log(db, user, "agent_reject", "agent_run", run.id, after={"reason": payload.reason}, reason=payload.reason or "人工拒绝 Agent 建议")
    db.commit()
    return {"ok": True, "run": agent_run_payload(run)}


@app.get("/api/media", response_model=dict)
def list_media(
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(current_user)],
    q: str | None = None,
    country: str | None = None,
    platform_type: str | None = None,
    min_volume: float | None = Query(default=None, ge=0),
    max_volume: float | None = Query(default=None, ge=0),
    cooperation_status: str | None = None,
    page: int = 1,
    page_size: int = 20,
):
    query = db.query(Media)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Media.name.ilike(like),
            Media.website_url.ilike(like),
            Media.profile_links.cast(String).ilike(like),
            Media.notes.ilike(like),
            Media.contacts.any(or_(Contact.name.ilike(like), Contact.email.ilike(like), Contact.phone.ilike(like))),
        ))
    if country:
        query = query.filter(Media.country == country)
    if platform_type:
        query = query.filter(Media.platform_type == platform_type)
    if min_volume is not None:
        query = query.filter(Media.followers_or_traffic >= min_volume)
    if max_volume is not None:
        query = query.filter(Media.followers_or_traffic <= max_volume)
    if cooperation_status:
        query = query.filter(Media.cooperation_status == cooperation_status)
    return list_payload(query.order_by(Media.updated_at.desc()), page, page_size)


def media_quality_report(db: Session) -> dict[str, Any]:
    items: list[dict[str, Any]] = []
    needs_review = 0
    all_media = db.query(Media).order_by(Media.updated_at.desc()).all()
    for item in all_media:
        changes: list[str] = []
        channel = "多平台" if len(item.profile_links or []) > 1 else normalize_channel(item.platform_type, item.website_url)
        cooperation = normalize_cooperation_status(item.cooperation_status)
        country, country_code, country_recognized = normalize_country(item.country)
        metric_type = infer_audience_metric_type(channel or item.platform_type)
        metric_k = metric_value_in_k(item.followers_or_traffic, item.audience_metric_unit)
        if channel and channel != item.platform_type:
            changes.append(f"渠道：{item.platform_type or '空白'} → {channel}")
        elif item.platform_type and not channel:
            needs_review += 1
        if cooperation and cooperation != item.cooperation_status:
            changes.append(f"合作状态：{item.cooperation_status or '空白'} → {cooperation}")
        elif item.cooperation_status and not cooperation:
            needs_review += 1
        if country and (country != item.country or country_code != item.country_code):
            changes.append(f"国家：{item.country or '空白'} → {country} ({country_code})")
        elif item.country and not country_recognized:
            needs_review += 1
        if item.audience_metric_type != metric_type:
            changes.append(f"指标：{item.audience_metric_type or '未标记'} → {metric_type}")
        if item.media_tier is not None:
            changes.append(f"停用等级：{item.media_tier} → 不再分级")
        if item.followers_or_traffic is not None and item.audience_metric_unit != "K":
            changes.append(f"粉丝/流量：{item.followers_or_traffic:g} → {metric_k:g} K")
        elif item.audience_metric_unit != "K":
            changes.append("单位：未标记 → K")
        if changes:
            items.append({"id": item.id, "name": item.name, "changes": changes})
    return {"total": len(all_media), "safe_changes": len(items), "needs_review": needs_review, "items": items}


@app.get("/api/media-data-quality")
def media_data_quality(db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(current_user)]):
    return media_quality_report(db)


@app.post("/api/media-data-quality/normalize")
def normalize_media_data(
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(editable_user)],
    change_reason: Annotated[str | None, Header(alias="X-Change-Reason")] = None,
):
    updated = 0
    needs_review = 0
    changed_ids: list[int] = []
    for item in db.query(Media).all():
        changed = False
        was_pending_status = item.cooperation_status == "待核验"
        channel = "多平台" if len(item.profile_links or []) > 1 else normalize_channel(item.platform_type, item.website_url)
        cooperation = normalize_cooperation_status(item.cooperation_status)
        country, country_code, country_recognized = normalize_country(item.country)
        metric_type = infer_audience_metric_type(channel or item.platform_type)
        metric_k = metric_value_in_k(item.followers_or_traffic, item.audience_metric_unit)
        if channel and channel != item.platform_type:
            item.platform_type = channel
            changed = True
        elif item.platform_type and not channel:
            needs_review += 1
        if cooperation and cooperation != item.cooperation_status:
            item.cooperation_status = cooperation
            changed = True
        elif item.cooperation_status and not cooperation:
            needs_review += 1
            item.verification_status = "有冲突"
        if country and country_recognized:
            if item.country != country or item.country_code != country_code:
                item.country = country
                item.country_code = country_code
                changed = True
        elif item.country:
            needs_review += 1
            item.verification_status = "有冲突"
            changed = True
        if was_pending_status:
            item.verification_status = "待核验"
            changed = True
        if item.audience_metric_type != metric_type:
            item.audience_metric_type = metric_type
            changed = True
        if item.media_tier is not None:
            item.media_tier = None
            changed = True
        if item.audience_metric_unit != "K":
            item.followers_or_traffic = metric_k
            item.audience_metric_unit = "K"
            changed = True
        if changed:
            updated += 1
            changed_ids.append(item.id)
    if changed_ids:
        add_audit_log(db, user, "normalize", "media", "batch", after={"updated_ids": changed_ids, "updated_count": updated}, reason=change_reason)
    db.commit()
    return {"updated": updated, "needs_review": needs_review}


def normalized_media_name(value: str | None) -> str:
    return re.sub(r"[^\w]+", "", (value or "").casefold(), flags=re.UNICODE)


def media_duplicate_candidates(db: Session) -> dict[int, list[dict[str, Any]]]:
    media = db.query(Media).order_by(Media.id).all()
    candidates: dict[int, list[dict[str, Any]]] = {}
    identities = {
        item.id: {identity for link in clean_profile_links(item.profile_links, item.website_url) if (identity := profile_identity(link.get("url")))}
        for item in media
    }
    for index, left in enumerate(media):
        left_name = normalized_media_name(left.name)
        for right in media[index + 1:]:
            right_name = normalized_media_name(right.name)
            shared_profiles = identities[left.id] & identities[right.id]
            similarity = SequenceMatcher(None, left_name, right_name).ratio() if left_name and right_name else 0
            same_country = bool(left.country and right.country and left.country == right.country)
            transposed_name = same_country and left_name[:4] == right_name[:4] and sorted(left_name) == sorted(right_name)
            name_match = left_name == right_name or transposed_name or (min(len(left_name), len(right_name)) >= 5 and abs(len(left_name) - len(right_name)) <= 2 and similarity >= 0.84 and same_country)
            if not shared_profiles and not name_match:
                continue
            reason = "主页完全相同" if shared_profiles else f"名称相似度 {round(similarity * 100)}%"
            for source, target in ((left, right), (right, left)):
                candidates.setdefault(source.id, []).append({
                    "id": target.id,
                    "name": target.name,
                    "country": target.country,
                    "platform_type": target.platform_type,
                    "reason": reason,
                })
    return candidates


def media_review_issues(item: Media, duplicate_items: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    has_contact_method = any(
        any((contact.email, contact.phone, contact.whatsapp, contact.telegram, contact.brief_email, contact.press_release_email))
        for contact in item.contacts
    )
    if not has_contact_method:
        issues.append({"code": "missing_contact", "category": "contact", "label": "缺少联系方式", "reason": "没有邮箱、电话、WhatsApp 或 Telegram，请补充可用联系入口。"})
    if item.verification_status != "已核验":
        for reason in re.findall(r"\[数据核验\]\s*([^\n]+)", item.notes or ""):
            issues.append({"code": "profile_data", "category": "profile", "label": "主页 / 指标异常", "reason": reason.strip()})
    if item.verification_status == "有冲突":
        issues.append({"code": "data_conflict", "category": "conflict", "label": "资料冲突", "reason": "标准字典无法可靠归一当前资料，请人工确认。"})
    if duplicate_items:
        issues.append({"code": "possible_duplicate", "category": "duplicate", "label": "疑似重复", "reason": "；".join(f"{candidate['name']}（{candidate['reason']}）" for candidate in duplicate_items), "candidates": duplicate_items})
    metric_exists = item.followers_or_traffic is not None
    if metric_exists and not item.metric_source:
        issues.append({"code": "missing_source", "category": "source", "label": "缺少数据来源", "reason": "粉丝量或网站流量没有记录来源，无法判断数据可信度。"})
    verified_at = item.metric_verified_at or item.last_verified_at
    if metric_exists and verified_at and verified_at < date.today() - timedelta(days=180):
        issues.append({"code": "stale_metric", "category": "stale", "label": "数据已过期", "reason": f"粉丝量或流量最后核验于 {verified_at}，建议重新确认。"})
    if item.data_capture_method == "agent" and item.data_confidence is not None and item.data_confidence < 0.8:
        issues.append({"code": "low_confidence", "category": "confidence", "label": "低置信度", "reason": f"Agent 置信度为 {round(item.data_confidence * 100)}%，需要人工核对。"})
    return issues


def media_review_rows(db: Session) -> list[dict[str, Any]]:
    rows = []
    duplicate_map = media_duplicate_candidates(db)
    candidates = db.query(Media).order_by(Media.updated_at.desc()).all()
    for item in candidates:
        if item.review_snoozed_until and item.review_snoozed_until >= date.today():
            continue
        issues = media_review_issues(item, duplicate_map.get(item.id))
        if not issues:
            continue
        rows.append({
            "id": item.id,
            "name": item.name,
            "country": item.country,
            "platform_type": item.platform_type,
            "website_url": item.website_url,
            "verification_status": item.verification_status,
            "issues": issues,
            "issue_codes": [issue["code"] for issue in issues],
            "categories": sorted({issue["category"] for issue in issues}),
            "priority": any(issue["category"] in {"duplicate", "contact", "profile", "conflict", "confidence"} for issue in issues),
            "review_reason": "；".join(issue["reason"] for issue in issues),
            "notes": item.notes,
            "snoozed_until": item.review_snoozed_until,
        })
    return rows


@app.get("/api/media-review-queue")
def media_review_queue(
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(current_user)],
):
    rows = media_review_rows(db)
    categories = ("duplicate", "contact", "profile", "conflict", "source", "stale", "confidence")
    category_counts = {category: sum(category in row["categories"] for row in rows) for category in categories}
    priority_total = sum(bool(row["priority"]) for row in rows)
    return {"items": rows, "total": priority_total, "all_total": len(rows), "maintenance_total": len(rows) - priority_total, "category_counts": category_counts}


@app.post("/api/media-review-queue/batch")
def batch_media_review(
    payload: MediaReviewBatchIn,
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(editable_user)],
):
    if payload.action not in {"resolve", "snooze"}:
        raise HTTPException(400, "Unsupported review action")
    items = db.query(Media).filter(Media.id.in_(set(payload.media_ids))).all()
    changed: list[int] = []
    skipped: list[dict[str, Any]] = []
    duplicate_map = media_duplicate_candidates(db)
    for item in items:
        issues = media_review_issues(item, duplicate_map.get(item.id))
        blocking_codes = {"missing_contact", "possible_duplicate", "missing_source", "stale_metric", "low_confidence", "data_conflict"}
        if payload.action == "resolve" and any(issue["code"] in blocking_codes for issue in issues):
            skipped.append({"id": item.id, "name": item.name, "reason": "仍有必须先修复的资料问题"})
            continue
        before = {"verification_status": item.verification_status, "review_snoozed_until": item.review_snoozed_until}
        if payload.action == "snooze":
            item.review_snoozed_until = date.today() + timedelta(days=payload.snooze_days)
        else:
            item.verification_status = "已核验"
            item.last_verified_at = date.today()
            item.review_snoozed_until = None
        add_audit_log(db, user, f"review_{payload.action}", "media", item.id, before=before, after={"verification_status": item.verification_status, "review_snoozed_until": item.review_snoozed_until}, reason="待核验中心批量处理")
        changed.append(item.id)
    db.commit()
    return {"changed": len(changed), "changed_ids": changed, "skipped": skipped}


@app.post("/api/media-review-queue/{media_id}/resolve")
def resolve_media_review(
    media_id: int,
    payload: MediaReviewResolveIn,
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(editable_user)],
    change_reason: Annotated[str | None, Header(alias="X-Change-Reason")] = None,
):
    item = db.get(Media, media_id)
    if not item:
        raise HTTPException(404, "Media not found")
    issues = media_review_issues(item, media_duplicate_candidates(db).get(item.id))
    if any(issue["code"] == "missing_contact" for issue in issues):
        raise HTTPException(400, "请先为该媒体补充可用联系方式")
    if payload.cooperation_status is not None and payload.cooperation_status not in COOPERATION_STATUSES:
        raise HTTPException(400, "请选择标准合作状态")
    product = db.get(Product, payload.product_id) if payload.product_id else None
    if payload.product_id and not product:
        raise HTTPException(404, "Product not found")
    if not product and (payload.product_name or "").strip():
        product = find_or_create_product(db, payload.product_name.strip(), "待核验中心确认")
    project = db.get(Project, payload.project_id) if payload.project_id else None
    if payload.project_id and not project:
        raise HTTPException(404, "Project not found")
    campaign = None
    if payload.create_collaboration:
        if not project:
            raise HTTPException(400, "创建执行单时必须选择项目")
        campaign = db.query(Campaign).filter(Campaign.project_id == project.id, Campaign.media_id == item.id, Campaign.archived_at.is_(None)).first()
        if not campaign:
            campaign = Campaign(
                project_id=project.id,
                media_id=item.id,
                product_id=product.id if product else None,
                collaboration_type=(payload.collaboration_type or "").strip() or None,
                execution_status="待确认",
                stage="Not Started",
                notes="由待核验中心建立",
            )
            db.add(campaign)
            db.flush()
            db.add(CampaignStageEvent(campaign_id=campaign.id, user_id=user.id, from_status=None, to_status="待确认", action="create", reason="待核验中心建立"))
        else:
            if product and not campaign.product_id:
                campaign.product_id = product.id
            if payload.collaboration_type and not campaign.collaboration_type:
                campaign.collaboration_type = payload.collaboration_type.strip()
        ensure_project_link(db, project.id, product.id) if product else None
    before = {"cooperation_status": item.cooperation_status, "verification_status": item.verification_status, "notes": item.notes}
    if payload.cooperation_status is not None:
        item.cooperation_status = payload.cooperation_status
    item.verification_status = "已核验"
    item.last_verified_at = date.today()
    item.review_snoozed_until = None
    result_note = f"[核验结果] {datetime.now().strftime('%Y-%m-%d')} 已完成人工核验"
    if payload.cooperation_status:
        result_note += f"；状态={payload.cooperation_status}"
    if product:
        result_note += f"；产品={product.model}"
    if project:
        result_note += f"；项目={project.name}"
    item.notes = "\n".join(part for part in [item.notes, result_note] if part)
    add_audit_log(db, user, "resolve_review", "media", item.id, before=before, after={"cooperation_status": item.cooperation_status, "product_id": product.id if product else None, "project_id": project.id if project else None, "campaign_id": campaign.id if campaign else None}, reason=change_reason)
    db.commit()
    return {"ok": True, "media_id": item.id, "campaign_id": campaign.id if campaign else None, "remaining": len(media_review_rows(db))}


def media_identity_matches(db: Session, links: list[dict], exclude_id: int | None = None) -> list[dict[str, Any]]:
    requested = {identity for link in links if (identity := profile_identity(link.get("url")))}
    if not requested:
        return []
    matches: list[dict[str, Any]] = []
    for candidate in db.query(Media).all():
        if candidate.id == exclude_id:
            continue
        candidate_links = clean_profile_links(candidate.profile_links, candidate.website_url)
        overlap = requested & {identity for link in candidate_links if (identity := profile_identity(link.get("url")))}
        if overlap:
            matches.append({"id": candidate.id, "name": candidate.name, "matched_profiles": sorted(overlap)})
    return matches


def require_unique_media_identity(db: Session, links: list[dict], exclude_id: int | None = None) -> None:
    matches = media_identity_matches(db, links, exclude_id)
    if matches:
        raise HTTPException(409, detail={"message": "主页已归属于其他媒体，请合并或修改主页", "matches": matches})


@app.get("/api/media-duplicates")
def media_duplicates(
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(current_user)],
    name: str,
    website_url: str | None = None,
    country: str | None = None,
):
    links = clean_profile_links(None, website_url)
    identity_items = media_identity_matches(db, links)
    if identity_items:
        ids = [item["id"] for item in identity_items]
        records = db.query(Media).filter(Media.id.in_(ids)).all()
        return {"items": jsonable_encoder(records), "reasons": identity_items}
    query = db.query(Media).filter(func.lower(Media.name) == name.strip().lower())
    if country:
        canonical_country, _, _ = normalize_country(country)
        query = query.filter(Media.country == (canonical_country or country.strip()))
    return {"items": jsonable_encoder(query.order_by(Media.updated_at.desc()).limit(10).all()), "reasons": []}


@app.post("/api/media", response_model=MediaOut)
def create_media(payload: MediaBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)], change_reason: Annotated[str | None, Header(alias="X-Change-Reason")] = None):
    normalized = normalize_media_payload(payload.model_dump())
    normalized["data_capture_method"] = normalized.get("data_capture_method") or "manual"
    normalized["data_source"] = normalized.get("data_source") or "CRM 人工录入"
    normalized["profile_links"] = clean_profile_links(normalized.get("profile_links"), normalized.get("website_url"))
    require_unique_media_identity(db, normalized["profile_links"])
    if len(normalized["profile_links"]) > 1:
        normalized["platform_type"] = "多平台"
    if normalized["profile_links"]:
        normalized["website_url"] = normalized["profile_links"][0]["url"]
    item = Media(**normalized)
    db.add(item)
    db.flush()
    add_audit_log(db, user, "create", "media", item.id, after=normalized, reason=change_reason)
    db.commit()
    db.refresh(item)
    return item


@app.get("/api/media/{item_id}")
def media_detail(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(current_user)]):
    item = db.query(Media).options(
        joinedload(Media.contacts),
        joinedload(Media.shipping_addresses).joinedload(ShippingAddress.contact),
        joinedload(Media.campaigns).joinedload(Campaign.product),
        joinedload(Media.campaigns).joinedload(Campaign.project),
        joinedload(Media.campaigns).joinedload(Campaign.owner),
        joinedload(Media.campaigns).joinedload(Campaign.activities).joinedload(Activity.user),
    ).filter(Media.id == item_id).first()
    if not item:
        raise HTTPException(404, "Media not found")
    deliverables = (
        db.query(Deliverable)
        .join(Campaign)
        .filter(Campaign.media_id == item_id)
        .order_by(Deliverable.published_at.desc().nullslast())
        .all()
    )
    return {
        "media": jsonable_encoder({
            "id": item.id,
            "name": item.name,
            "country": item.country,
            "country_code": item.country_code,
            "region": item.region,
            "category": item.category,
            "platform_type": item.platform_type,
            "website_url": item.website_url,
            "profile_links": item.profile_links or clean_profile_links(None, item.website_url),
            "followers_or_traffic": item.followers_or_traffic,
            "audience_metric_type": item.audience_metric_type,
            "audience_metric_unit": item.audience_metric_unit,
            "metric_source": item.metric_source,
            "metric_verified_at": item.metric_verified_at,
            "media_tier": item.media_tier,
            "cooperation_status": item.cooperation_status,
            "verification_status": item.verification_status,
            "notes": item.notes,
        }),
        "contacts": [jsonable_encoder({
            "id": contact.id,
            "media_id": contact.media_id,
            "name": contact.name,
            "role": contact.role,
            "email": contact.email,
            "phone": contact.phone,
            "whatsapp": contact.whatsapp,
            "telegram": contact.telegram,
            "brief_email": contact.brief_email,
            "press_release_email": contact.press_release_email,
            "is_primary": contact.is_primary,
            "notes": contact.notes,
        }) for contact in item.contacts],
        "shipping_addresses": [ShippingAddressOut.model_validate(address).model_dump(mode="json") for address in sorted(item.shipping_addresses, key=lambda address: (not address.is_default, address.id))],
        "products": list({campaign.product.id: {"id": campaign.product.id, "model": campaign.product.model, "full_name": campaign.product.full_name} for campaign in item.campaigns if campaign.product}.values()),
        "campaigns": [jsonable_encoder({
            "id": campaign.id,
            "project": {"id": campaign.project.id, "name": campaign.project.name} if campaign.project else None,
            "product": {"id": campaign.product.id, "model": campaign.product.model} if campaign.product else None,
            "owner": {"id": campaign.owner.id, "name": campaign.owner.name} if campaign.owner else None,
            "collaboration_type": campaign.collaboration_type,
            "execution_status": campaign.execution_status,
            "next_action": campaign.next_action,
            "follow_up_date": campaign.follow_up_date,
            "expected_publish_date": campaign.expected_publish_date,
            "updated_at": campaign.updated_at,
            "activities": [{"id": activity.id, "activity_type": activity.activity_type, "content": activity.content, "created_at": activity.created_at, "user": activity.user.name if activity.user else None} for activity in sorted(campaign.activities, key=lambda row: row.created_at, reverse=True)],
        }) for campaign in sorted(item.campaigns, key=lambda row: row.updated_at, reverse=True)],
        "deliverables": [{"id": deliverable.id, "url": deliverable.url, "deliverable_type": deliverable.deliverable_type, "published_at": deliverable.published_at} for deliverable in deliverables],
    }


@app.put("/api/media/{item_id}", response_model=MediaOut)
def update_media(item_id: int, payload: MediaBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)], change_reason: Annotated[str | None, Header(alias="X-Change-Reason")] = None):
    item = db.get(Media, item_id)
    if not item:
        raise HTTPException(404, "Media not found")
    before = MediaOut.model_validate(item).model_dump(mode="json")
    normalized = normalize_media_payload(payload.model_dump())
    normalized["profile_links"] = clean_profile_links(normalized.get("profile_links"), normalized.get("website_url"))
    require_unique_media_identity(db, normalized["profile_links"], item_id)
    if len(normalized["profile_links"]) > 1:
        normalized["platform_type"] = "多平台"
    normalized["website_url"] = normalized["profile_links"][0]["url"] if normalized["profile_links"] else None
    for key, value in normalized.items():
        setattr(item, key, value)
    add_audit_log(db, user, "update", "media", item.id, before=before, after=normalized, reason=change_reason)
    db.commit()
    db.refresh(item)
    return item


@app.delete("/api/media/{item_id}")
def delete_media(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))]):
    item = db.get(Media, item_id)
    if not item:
        raise HTTPException(404, "Media not found")
    counts = {
        "campaigns": db.query(func.count(Campaign.id)).filter(Campaign.media_id == item_id).scalar() or 0,
        "contacts": db.query(func.count(Contact.id)).filter(Contact.media_id == item_id).scalar() or 0,
        "addresses": db.query(func.count(ShippingAddress.id)).filter(ShippingAddress.media_id == item_id).scalar() or 0,
    }
    if any(counts.values()):
        raise HTTPException(409, detail={"message": "该媒体仍有关联数据，请先合并到正确媒体，避免丢失合作历史。", "counts": counts})
    before = MediaOut.model_validate(item).model_dump(mode="json")
    add_audit_log(db, user, "delete", "media", item.id, before=before, reason="管理员永久删除无关联媒体")
    db.delete(item)
    db.commit()
    return {"ok": True}


@app.post("/api/media/{item_id}/merge")
def merge_media(item_id: int, payload: MediaMergeIn, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))]):
    source = db.query(Media).options(joinedload(Media.contacts), joinedload(Media.shipping_addresses), joinedload(Media.campaigns)).filter(Media.id == item_id).first()
    target = db.query(Media).options(joinedload(Media.contacts), joinedload(Media.shipping_addresses), joinedload(Media.campaigns)).filter(Media.id == payload.target_media_id).first()
    if not source or not target:
        raise HTTPException(404, "Media not found")
    if source.id == target.id:
        raise HTTPException(400, "请选择不同的目标媒体")
    before = {
        "source": MediaOut.model_validate(source).model_dump(mode="json"),
        "target": MediaOut.model_validate(target).model_dump(mode="json"),
        "campaign_ids": [campaign.id for campaign in source.campaigns],
        "contact_ids": [contact.id for contact in source.contacts],
        "address_ids": [address.id for address in source.shipping_addresses],
    }
    moved = {"campaigns": 0, "contacts": 0, "addresses": 0, "duplicate_contacts": 0}
    target_contacts_by_email = {contact.email.strip().lower(): contact for contact in target.contacts if contact.email}
    for contact in list(source.contacts):
        duplicate = target_contacts_by_email.get(contact.email.strip().lower()) if contact.email else None
        if duplicate:
            for address in list(contact.shipping_addresses):
                address.contact = duplicate
            for field in ("name", "role", "phone", "whatsapp", "telegram", "brief_email", "press_release_email", "notes"):
                if not getattr(duplicate, field) and getattr(contact, field):
                    setattr(duplicate, field, getattr(contact, field))
            db.delete(contact)
            moved["duplicate_contacts"] += 1
        else:
            contact.media = target
            if contact.email:
                target_contacts_by_email[contact.email.strip().lower()] = contact
            moved["contacts"] += 1
    target_has_default = any(address.is_default for address in target.shipping_addresses)
    for address in list(source.shipping_addresses):
        address.media = target
        if target_has_default and address.is_default:
            address.is_default = False
        elif address.is_default:
            target_has_default = True
        moved["addresses"] += 1
    for campaign in list(source.campaigns):
        campaign.media = target
        moved["campaigns"] += 1
    target.profile_links = clean_profile_links([*(target.profile_links or []), *(source.profile_links or [])], target.website_url or source.website_url)
    if target.profile_links:
        target.website_url = target.profile_links[0]["url"]
    for field in ("country", "country_code", "region", "category", "platform_type", "followers_or_traffic", "audience_metric_type", "audience_metric_unit", "metric_source", "metric_verified_at"):
        if getattr(target, field) in (None, "") and getattr(source, field) not in (None, ""):
            setattr(target, field, getattr(source, field))
    cooperation_rank = {"未联系": 0, "待回复": 1, "洽谈中": 2, "已合作": 3, "暂缓": 1, "不合作": 1}
    if cooperation_rank.get(source.cooperation_status or "", 0) > cooperation_rank.get(target.cooperation_status or "", 0):
        target.cooperation_status = source.cooperation_status
    db.flush()
    db.delete(source)
    add_audit_log(db, user, "merge", "media", target.id, before=before, after={"target_media_id": target.id, "source_media_id": item_id, **moved}, reason=f"合并重复媒体 {source.name} → {target.name}")
    db.commit()
    return {"ok": True, "source_media_id": item_id, "target_media_id": target.id, **moved}


@app.get("/api/media/{item_id}/merge-preview")
def preview_media_merge(
    item_id: int,
    target_media_id: int,
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(require_roles("Admin"))],
):
    source = db.query(Media).options(joinedload(Media.contacts), joinedload(Media.shipping_addresses), joinedload(Media.campaigns)).filter(Media.id == item_id).first()
    target = db.query(Media).options(joinedload(Media.contacts), joinedload(Media.shipping_addresses), joinedload(Media.campaigns)).filter(Media.id == target_media_id).first()
    if not source or not target:
        raise HTTPException(404, "Media not found")
    if source.id == target.id:
        raise HTTPException(400, "请选择不同的目标媒体")
    target_emails = {normalized_email(contact.email) for contact in target.contacts if normalized_email(contact.email)}
    duplicate_contacts = [contact for contact in source.contacts if normalized_email(contact.email) in target_emails]
    source_data = MediaOut.model_validate(source).model_dump(mode="json")
    target_data = MediaOut.model_validate(target).model_dump(mode="json")
    differing_fields = []
    for field in ("name", "country", "platform_type", "followers_or_traffic", "cooperation_status", "metric_source", "metric_verified_at"):
        if source_data.get(field) not in (None, "") and target_data.get(field) not in (None, "") and source_data.get(field) != target_data.get(field):
            differing_fields.append({"field": field, "source": source_data.get(field), "target": target_data.get(field), "kept": "target"})
    return {
        "source": source_data,
        "target": target_data,
        "moves": {
            "campaigns": len(source.campaigns),
            "contacts": len(source.contacts) - len(duplicate_contacts),
            "duplicate_contacts": len(duplicate_contacts),
            "addresses": len(source.shipping_addresses),
            "profile_links": len(clean_profile_links(source.profile_links, source.website_url)),
        },
        "field_conflicts": differing_fields,
    }


@app.get("/api/products", response_model=dict)
def list_products(db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(current_user)], q: str | None = None, page: int = 1, page_size: int = 20):
    query = db.query(Product).options(joinedload(Product.project_links).joinedload(ProjectProduct.project))
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Product.model.ilike(like), Product.full_name.ilike(like), Product.aliases.ilike(like)))
    total = query.count()
    items = query.order_by(Product.model).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": [product_payload(db, item) for item in items], "total": total}


def product_payload(db: Session, item: Product) -> dict:
    payload = jsonable_encoder(item)
    payload["project_ids"] = [link.project_id for link in item.project_links]
    payload["projects"] = [{"id": link.project.id, "name": link.project.name, "project_code": link.project.project_code} for link in item.project_links if link.project]
    payload["shipment_count"] = db.query(func.count(ShipmentItem.id)).filter(ShipmentItem.product_id == item.id).scalar() or 0
    return payload


def normalized_alias_text(value: str | None) -> str | None:
    aliases: list[str] = []
    seen: set[str] = set()
    for raw in re.split(r"[,，;；\n|/]+", value or ""):
        label = raw.strip()
        identity = product_identity(label)
        if label and identity and identity not in seen:
            seen.add(identity)
            aliases.append(label)
    return ", ".join(aliases) or None


def require_unique_product(db: Session, data: dict[str, Any], exclude_id: int | None = None) -> None:
    requested = {
        product_identity(value)
        for value in [data.get("model"), data.get("full_name"), *re.split(r"[,，;；\n|/]+", data.get("aliases") or "")]
        if product_identity(value)
    }
    matches = []
    for item in db.query(Product).all():
        if item.id != exclude_id and requested & product_aliases(item):
            matches.append({"id": item.id, "model": item.model, "aliases": sorted(requested & product_aliases(item))})
    if matches:
        raise HTTPException(409, detail={"message": "产品型号或别名已被占用，请合并产品或调整别名", "matches": matches})


def sync_product_projects(db: Session, item: Product, project_ids: list[int]) -> None:
    valid_ids = {project.id for project in db.query(Project).filter(Project.id.in_(project_ids)).all()} if project_ids else set()
    if len(valid_ids) != len(set(project_ids)):
        raise HTTPException(400, "One or more projects do not exist")
    for link in list(item.project_links):
        if link.project_id not in valid_ids:
            db.delete(link)
    current_ids = {link.project_id for link in item.project_links}
    for project_id in valid_ids - current_ids:
        db.add(ProjectProduct(project_id=project_id, product_id=item.id))


@app.post("/api/products")
def create_product(payload: ProductBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    data = payload.model_dump()
    project_ids = data.pop("project_ids", [])
    data["model"] = data["model"].strip()
    data["aliases"] = normalized_alias_text(data.get("aliases"))
    require_unique_product(db, data)
    if not data.get("platform"):
        from .product_backfill import chipset_from_model
        data["platform"] = chipset_from_model(data["model"])
    item = Product(**data)
    db.add(item)
    db.flush()
    sync_product_projects(db, item, project_ids)
    db.commit()
    db.refresh(item)
    return product_payload(db, item)


@app.get("/api/products/{item_id}")
def product_detail(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(current_user)]):
    item = db.get(Product, item_id)
    if not item:
        raise HTTPException(404, "Product not found")
    campaigns = (
        db.query(Campaign)
        .options(joinedload(Campaign.media), joinedload(Campaign.owner), joinedload(Campaign.deliverables))
        .filter(Campaign.product_id == item_id)
        .order_by(Campaign.updated_at.desc())
        .all()
    )
    db.refresh(item)
    return {"product": product_payload(db, item), "campaigns": campaigns}


@app.put("/api/products/{item_id}")
def update_product(item_id: int, payload: ProductBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    item = db.get(Product, item_id)
    if not item:
        raise HTTPException(404, "Product not found")
    data = payload.model_dump()
    project_ids = data.pop("project_ids", [])
    data["model"] = data["model"].strip()
    data["aliases"] = normalized_alias_text(data.get("aliases"))
    require_unique_product(db, data, item_id)
    for key, value in data.items():
        setattr(item, key, value)
    sync_product_projects(db, item, project_ids)
    db.commit()
    db.refresh(item)
    return product_payload(db, item)


@app.delete("/api/products/{item_id}")
def delete_product(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))]):
    item = db.get(Product, item_id)
    if not item:
        raise HTTPException(404, "Product not found")
    campaign_count = db.query(func.count(Campaign.id)).filter(Campaign.product_id == item_id).scalar() or 0
    shipment_count = db.query(func.count(ShipmentItem.id)).filter(ShipmentItem.product_id == item_id).scalar() or 0
    if campaign_count or shipment_count:
        raise HTTPException(400, "This product is still referenced. Merge it into another product before deleting.")
    db.delete(item)
    db.commit()
    return {"ok": True}


@app.post("/api/products/{item_id}/merge")
def merge_product(item_id: int, payload: ProductMergeIn, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))]):
    source = db.query(Product).options(joinedload(Product.project_links)).filter(Product.id == item_id).first()
    target = db.query(Product).options(joinedload(Product.project_links)).filter(Product.id == payload.target_product_id).first()
    if not source or not target:
        raise HTTPException(404, "Product not found")
    if source.id == target.id:
        raise HTTPException(400, "Choose a different target product")
    campaign_count = db.query(func.count(Campaign.id)).filter(Campaign.product_id == source.id).scalar() or 0
    shipment_count = db.query(func.count(ShipmentItem.id)).filter(ShipmentItem.product_id == source.id).scalar() or 0
    target_project_ids = {link.project_id for link in target.project_links}
    for link in source.project_links:
        if link.project_id not in target_project_ids:
            db.add(ProjectProduct(project_id=link.project_id, product_id=target.id))
    retained_aliases = [target.aliases, source.model, source.full_name, source.aliases]
    target.aliases = normalized_alias_text(", ".join(value for value in retained_aliases if value))
    db.query(Campaign).filter(Campaign.product_id == source.id).update({Campaign.product_id: target.id}, synchronize_session=False)
    db.query(ShipmentItem).filter(ShipmentItem.product_id == source.id).update({ShipmentItem.product_id: target.id, ShipmentItem.product_name: target.model}, synchronize_session=False)
    db.delete(source)
    db.commit()
    return {"ok": True, "campaign_count": campaign_count, "shipment_count": shipment_count, "target_product_id": target.id}


@app.get("/api/projects", response_model=dict)
def list_projects(db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(current_user)], q: str | None = None, status: str | None = None, history_only: bool = False, page: int = 1, page_size: int = 100):
    query = db.query(Project).options(joinedload(Project.owner))
    if history_only:
        query = query.filter(or_(Project.is_archived.is_(True), Project.name.like(f"{HISTORICAL_PROJECT_PREFIX}%")))
    else:
        query = query.filter(Project.is_archived.is_(False), ~Project.name.like(f"{HISTORICAL_PROJECT_PREFIX}%"))
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Project.name.ilike(like), Project.project_code.ilike(like), Project.objective.ilike(like)))
    if status:
        query = query.filter(Project.status == status)
    items = query.order_by(Project.updated_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    rows = []
    for item in items:
        actual = db.query(func.coalesce(func.sum(CostItem.actual_amount), 0)).join(Campaign).filter(Campaign.project_id == item.id).scalar() or 0
        rows.append({**jsonable_encoder(item), "owner": jsonable_encoder(item.owner) if item.owner else None, "actual_amount": actual, "campaign_count": len(item.campaigns)})
    return {"items": rows, "total": query.count()}


@app.post("/api/projects", response_model=ProjectOut)
def create_project(payload: ProjectBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    if payload.project_code and db.query(Project).filter(Project.project_code == payload.project_code).first():
        raise HTTPException(400, "Project code already exists")
    item = Project(**payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@app.get("/api/projects/{item_id}")
def project_detail(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(current_user)]):
    item = db.query(Project).options(joinedload(Project.owner), joinedload(Project.product_links).joinedload(ProjectProduct.product), joinedload(Project.campaigns).joinedload(Campaign.media), joinedload(Project.campaigns).joinedload(Campaign.owner), joinedload(Project.campaigns).joinedload(Campaign.shipments).joinedload(Shipment.items), joinedload(Project.campaigns).joinedload(Campaign.deliverables), joinedload(Project.campaigns).joinedload(Campaign.cost_items), joinedload(Project.campaigns).joinedload(Campaign.activities)).filter(Project.id == item_id).first()
    if not item:
        raise HTTPException(404, "Project not found")
    planned = sum(cost.planned_amount or 0 for campaign in item.campaigns for cost in campaign.cost_items)
    actual = sum(cost.actual_amount or 0 for campaign in item.campaigns for cost in campaign.cost_items)
    return jsonable_encoder({
        "project": project_detail_payload(item),
        "planned_amount": planned,
        "actual_amount": actual,
        "summary": project_result_summary(item),
        "campaigns": [campaign_detail_payload(campaign) for campaign in item.campaigns],
        "products": [product_detail_payload(link.product) for link in item.product_links if link.product],
    })


def user_summary_payload(item: User | None) -> dict | None:
    if not item:
        return None
    return {"id": item.id, "name": item.name, "email": item.email, "role": item.role}


def media_summary_payload(item: Media | None) -> dict | None:
    if not item:
        return None
    return {"id": item.id, "name": item.name, "country": item.country, "platform_type": item.platform_type, "website_url": item.website_url}


def product_detail_payload(item: Product) -> dict:
    return {
        "id": item.id,
        "model": item.model,
        "full_name": item.full_name,
        "product_line": item.product_line,
        "platform": item.platform,
        "aliases": item.aliases,
        "launch_status": item.launch_status,
        "notes": item.notes,
    }


def deliverable_detail_payload(item: Deliverable) -> dict[str, Any]:
    snapshots = sorted(item.performance_snapshots or [], key=lambda row: row.captured_at)
    return {
        "id": item.id,
        "campaign_id": item.campaign_id,
        "deliverable_type": item.deliverable_type,
        "title": item.title,
        "url": item.url,
        "published_at": item.published_at,
        "views": item.views,
        "likes": item.likes,
        "comments": item.comments,
        "impressions": item.impressions,
        "data_updated_at": item.data_updated_at,
        "performance_notes": item.performance_notes,
        "platform_content_id": item.platform_content_id,
        "platform_channel_id": item.platform_channel_id,
        "matched_tag": item.matched_tag,
        "match_method": item.match_method,
        "platform_published_at": item.platform_published_at,
        "first_detected_at": item.first_detected_at,
        "monitoring_status": item.monitoring_status,
        "monitoring_completed_at": item.monitoring_completed_at,
        "performance_snapshots": [{
            "id": row.id,
            "sample_kind": row.sample_kind,
            "views": row.views,
            "likes": row.likes,
            "comments": row.comments,
            "captured_at": row.captured_at,
            "hours_since_publish": row.hours_since_publish,
            "source": row.source,
        } for row in snapshots],
    }


def project_detail_payload(item: Project) -> dict:
    return {
        "id": item.id,
        "name": item.name,
        "project_code": item.project_code,
        "owner_id": item.owner_id,
        "owner": user_summary_payload(item.owner),
        "objective": item.objective,
        "status": item.status,
        "start_date": item.start_date,
        "end_date": item.end_date,
        "budget_amount": item.budget_amount,
        "budget_currency": item.budget_currency,
        "notes": item.notes,
        "is_archived": item.is_archived,
        "archived_at": item.archived_at,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }


def shipment_detail_payload(item: Shipment) -> dict:
    return {
        "id": item.id,
        "shipping_address_id": item.shipping_address_id,
        "recipient_address": item.recipient_address,
        "oa_pi_number": item.oa_pi_number,
        "tracking_number": item.tracking_number,
        "carrier": item.carrier,
        "status": item.status,
        "shipped_at": item.shipped_at,
        "delivered_at": item.delivered_at,
        "notes": item.notes,
        "items": [{"id": row.id, "product_id": row.product_id, "product_name": row.product_name, "quantity": row.quantity, "unit_cost": row.unit_cost} for row in item.items],
    }


def validate_shipping_address_links(db: Session, media_id: int, contact_id: int | None) -> None:
    if not db.get(Media, media_id):
        raise HTTPException(404, "Media not found")
    if contact_id:
        contact = db.get(Contact, contact_id)
        if not contact or contact.media_id != media_id:
            raise HTTPException(400, "Contact does not belong to the selected media")


CONTACT_MERGE_FIELDS = ["name", "role", "email", "phone", "whatsapp", "telegram", "brief_email", "press_release_email", "notes"]


def normalized_contact_value(value: str | None) -> str:
    return re.sub(r"[^\w@]+", "", (value or "").strip().lower(), flags=re.UNICODE)


def exact_contact_duplicate_groups(db: Session) -> list[list[Contact]]:
    grouped: dict[tuple[int, str, str, str], list[Contact]] = {}
    for contact in db.query(Contact).order_by(Contact.id).all():
        key = (
            contact.media_id,
            normalized_contact_value(contact.name),
            normalized_contact_value(contact.email),
            normalized_contact_value(contact.phone),
        )
        if not any(key[1:]):
            continue
        grouped.setdefault(key, []).append(contact)
    return [items for items in grouped.values() if len(items) > 1]


def contact_duplicate_report(db: Session) -> dict[str, Any]:
    groups = exact_contact_duplicate_groups(db)
    duplicate_ids = [contact.id for group in groups for contact in group[1:]]
    address_count = db.query(ShippingAddress).filter(ShippingAddress.contact_id.in_(duplicate_ids)).count() if duplicate_ids else 0
    return {
        "contact_total": db.query(Contact).count(),
        "duplicate_groups": len(groups),
        "duplicate_rows": len(duplicate_ids),
        "linked_addresses": address_count,
        "items": [{
            "media_id": group[0].media_id,
            "name": group[0].name,
            "count": len(group),
            "contact_ids": [contact.id for contact in group],
        } for group in groups],
    }


@app.get("/api/contact-duplicates")
def contact_duplicates(db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(current_user)]):
    return contact_duplicate_report(db)


@app.post("/api/contact-duplicates/merge")
def merge_contact_duplicates(
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(editable_user)],
    change_reason: Annotated[str | None, Header(alias="X-Change-Reason")] = None,
):
    groups = exact_contact_duplicate_groups(db)
    removed_ids: list[int] = []
    kept_ids: list[int] = []
    transferred_addresses = 0
    for group in groups:
        def score(contact: Contact) -> tuple[int, int, int]:
            populated = sum(bool(getattr(contact, field, None)) for field in CONTACT_MERGE_FIELDS)
            return (1 if contact.is_primary else 0, populated, -contact.id)

        master = max(group, key=score)
        duplicates = [contact for contact in group if contact.id != master.id]
        kept_ids.append(master.id)
        master.is_primary = any(contact.is_primary for contact in group)
        merged_notes = []
        for contact in group:
            if contact.notes and contact.notes.strip() not in merged_notes:
                merged_notes.append(contact.notes.strip())
        for field in CONTACT_MERGE_FIELDS:
            if field == "notes":
                continue
            if not getattr(master, field, None):
                replacement = next((getattr(contact, field) for contact in group if getattr(contact, field, None)), None)
                if replacement:
                    setattr(master, field, replacement)
        if merged_notes:
            master.notes = "\n\n".join(merged_notes)
        for duplicate in duplicates:
            transferred_addresses += db.query(ShippingAddress).filter(ShippingAddress.contact_id == duplicate.id).update({ShippingAddress.contact_id: master.id}, synchronize_session=False)
            removed_ids.append(duplicate.id)
            db.delete(duplicate)
    if removed_ids:
        add_audit_log(db, user, "merge_duplicates", "contact", "batch", before={"removed_ids": removed_ids}, after={"kept_ids": kept_ids, "removed_count": len(removed_ids), "transferred_addresses": transferred_addresses}, reason=change_reason)
    db.commit()
    return {"merged_groups": len(groups), "removed": len(removed_ids), "remaining": db.query(Contact).count(), "transferred_addresses": transferred_addresses}


def set_default_shipping_address(db: Session, item: ShippingAddress) -> None:
    db.query(ShippingAddress).filter(
        ShippingAddress.media_id == item.media_id,
        ShippingAddress.id != item.id,
    ).update({ShippingAddress.is_default: False}, synchronize_session=False)
    item.is_default = True


def format_shipping_address(item: ShippingAddress) -> str:
    parts = [
        item.recipient_name,
        item.phone,
        item.email,
        item.address_text,
        " ".join(value for value in [item.city, item.region, item.postal_code, item.country] if value),
        f"税号/清关号: {item.tax_or_customs_number}" if item.tax_or_customs_number else None,
        item.shipping_notes,
    ]
    return "\n".join(str(value).strip() for value in parts if value and str(value).strip())


def apply_shipping_address_snapshot(db: Session, media_id: int, data: dict[str, Any]) -> None:
    address_id = data.get("shipping_address_id")
    if not address_id:
        return
    address = db.get(ShippingAddress, address_id)
    if not address or address.media_id != media_id:
        raise HTTPException(400, "Shipping address does not belong to the selected media")
    if not data.get("recipient_address"):
        data["recipient_address"] = format_shipping_address(address)


def campaign_detail_payload(item: Campaign) -> dict:
    return {
        "id": item.id,
        "project_id": item.project_id,
        "media_id": item.media_id,
        "owner_id": item.owner_id,
        "collaboration_type": item.collaboration_type,
        "execution_status": item.execution_status,
        "expected_publish_date": item.expected_publish_date,
        "next_action": item.next_action,
        "follow_up_date": item.follow_up_date,
        "follow_up_priority": item.follow_up_priority,
        "follow_up_done": item.follow_up_done,
        "notes": item.notes,
        "media": media_summary_payload(item.media),
        "owner": user_summary_payload(item.owner),
        "shipments": [shipment_detail_payload(shipment) for shipment in item.shipments],
        "deliverables": [deliverable_detail_payload(row) for row in item.deliverables],
        "cost_items": [{"id": row.id, "cost_type": row.cost_type, "planned_amount": row.planned_amount, "actual_amount": row.actual_amount, "currency": row.currency, "payment_status": row.payment_status} for row in item.cost_items],
        "activities": [{"id": row.id, "activity_type": row.activity_type, "content": row.content, "created_at": row.created_at} for row in item.activities],
    }


def project_result_summary(item: Project) -> dict[str, Any]:
    campaigns = item.campaigns
    deliverables = [deliverable for campaign in campaigns for deliverable in campaign.deliverables]
    actual = sum(cost.actual_amount or 0 for campaign in campaigns for cost in campaign.cost_items)
    impressions = sum(deliverable.impressions or 0 for deliverable in deliverables)
    views = sum(deliverable.views or 0 for deliverable in deliverables)
    likes = sum(deliverable.likes or 0 for deliverable in deliverables)
    comments = sum(deliverable.comments or 0 for deliverable in deliverables)
    day_3_samples = [
        snapshot
        for deliverable in deliverables
        for snapshot in deliverable.performance_snapshots
        if snapshot.sample_kind == "day_3"
    ]
    three_day_views = sum(snapshot.views or 0 for snapshot in day_3_samples)
    three_day_likes = sum(snapshot.likes or 0 for snapshot in day_3_samples)
    three_day_comments = sum(snapshot.comments or 0 for snapshot in day_3_samples)
    reach = impressions or views
    published = sum(1 for campaign in campaigns if campaign.execution_status in {"已发布", "已结算"} or campaign.deliverables)
    return {
        "collaboration_count": len(campaigns),
        "published_count": published,
        "completion_rate": round(published / len(campaigns) * 100, 1) if campaigns else 0,
        "actual_amount": actual,
        "deliverable_count": len(deliverables),
        "impressions": impressions,
        "views": views,
        "likes": likes,
        "comments": comments,
        "three_day_content_count": len(day_3_samples),
        "three_day_views": three_day_views,
        "three_day_likes": three_day_likes,
        "three_day_comments": three_day_comments,
        "three_day_cpv": round(actual / three_day_views, 4) if three_day_views else None,
        "cpm_base": "曝光" if impressions else ("播放/阅读" if views else None),
        "cpm": round(actual / reach * 1000, 2) if reach else None,
    }


@app.get("/api/projects/{item_id}/summary")
def project_summary(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(current_user)]):
    item = db.query(Project).options(joinedload(Project.campaigns).joinedload(Campaign.media), joinedload(Project.campaigns).joinedload(Campaign.deliverables), joinedload(Project.campaigns).joinedload(Campaign.cost_items)).filter(Project.id == item_id).first()
    if not item:
        raise HTTPException(404, "Project not found")
    return project_result_summary(item)


@app.get("/api/projects/{item_id}/report.xlsx")
def export_project_report(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(current_user)]):
    item = db.query(Project).options(joinedload(Project.campaigns).joinedload(Campaign.media), joinedload(Project.campaigns).joinedload(Campaign.deliverables), joinedload(Project.campaigns).joinedload(Campaign.cost_items)).filter(Project.id == item_id).first()
    if not item:
        raise HTTPException(404, "Project not found")
    summary = project_result_summary(item)
    book = Workbook()
    overview = book.active
    overview.title = "项目成果摘要"
    overview.append(["项目", item.name])
    overview.append(["项目编号", item.project_code or ""])
    overview.append(["合作对象数", summary["collaboration_count"]])
    overview.append(["已发布", summary["published_count"]])
    overview.append(["完成率", summary["completion_rate"] / 100])
    overview.append(["实付", summary["actual_amount"]])
    overview.append(["总曝光", summary["impressions"]])
    overview.append(["总播放/阅读", summary["views"]])
    overview.append(["总互动", summary["likes"] + summary["comments"]])
    overview.append(["CPM", summary["cpm"] if summary["cpm"] is not None else "待补充曝光/播放数据"])
    overview["A1"].font = Font(bold=True, color="FFFFFF")
    overview["B1"].font = Font(bold=True, color="FFFFFF")
    overview["A1"].fill = PatternFill("solid", fgColor="1B8D82")
    overview["B1"].fill = PatternFill("solid", fgColor="1B8D82")
    overview.column_dimensions["A"].width = 22
    overview.column_dimensions["B"].width = 34
    overview["B5"].number_format = "0.0%"
    overview["B6"].number_format = '#,##0.00'
    overview["B10"].number_format = '#,##0.00'
    details = book.create_sheet("合作明细")
    details.append(["媒体/KOL", "执行状态", "内容链接", "发布时间", "曝光", "播放/阅读", "点赞", "评论", "实付"])
    for campaign in item.campaigns:
        cost = sum(row.actual_amount or 0 for row in campaign.cost_items)
        if campaign.deliverables:
            for deliverable in campaign.deliverables:
                details.append([campaign.media.name if campaign.media else "", campaign.execution_status, deliverable.url or "", deliverable.published_at, deliverable.impressions or 0, deliverable.views or 0, deliverable.likes or 0, deliverable.comments or 0, cost])
        else:
            details.append([campaign.media.name if campaign.media else "", campaign.execution_status, "", "", 0, 0, 0, 0, cost])
    for cell in details[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1B8D82")
    for column, width in {"A": 24, "B": 16, "C": 40, "D": 14, "E": 12, "F": 14, "G": 12, "H": 12, "I": 14}.items():
        details.column_dimensions[column].width = width
    stream = BytesIO()
    book.save(stream)
    stream.seek(0)
    filename = f"{item.name}_项目复盘.xlsx"
    disposition = f"attachment; filename=project-report.xlsx; filename*=UTF-8''{quote(filename)}"
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": disposition})


@app.put("/api/projects/{item_id}", response_model=ProjectOut)
def update_project(item_id: int, payload: ProjectBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    item = db.get(Project, item_id)
    if not item:
        raise HTTPException(404, "Project not found")
    if payload.project_code and db.query(Project).filter(Project.project_code == payload.project_code, Project.id != item_id).first():
        raise HTTPException(400, "Project code already exists")
    for key, value in payload.model_dump().items():
        setattr(item, key, value)
    db.commit()
    db.refresh(item)
    return item


@app.delete("/api/projects/{item_id}")
def delete_project(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))]):
    item = db.query(Project).options(joinedload(Project.campaigns).joinedload(Campaign.shipments), joinedload(Project.campaigns).joinedload(Campaign.cost_items), joinedload(Project.campaigns).joinedload(Campaign.deliverables), joinedload(Project.campaigns).joinedload(Campaign.activities)).filter(Project.id == item_id).first()
    if not item:
        raise HTTPException(404, "Project not found")
    deleted_campaigns = len(item.campaigns)
    deleted_shipments = sum(len(campaign.shipments) for campaign in item.campaigns)
    deleted_cost_items = sum(len(campaign.cost_items) for campaign in item.campaigns)
    deleted_deliverables = sum(len(campaign.deliverables) for campaign in item.campaigns)
    deleted_activities = sum(len(campaign.activities) for campaign in item.campaigns)
    for campaign in list(item.campaigns):
        db.delete(campaign)
    db.flush()
    db.delete(item)
    db.commit()
    return {"ok": True, "deleted_campaigns": deleted_campaigns, "deleted_shipments": deleted_shipments, "deleted_cost_items": deleted_cost_items, "deleted_deliverables": deleted_deliverables, "deleted_activities": deleted_activities}


@app.post("/api/projects/{item_id}/archive")
def archive_project(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))]):
    item = db.get(Project, item_id)
    if not item:
        raise HTTPException(404, "Project not found")
    item.is_archived = True
    item.archived_at = datetime.now()
    db.commit()
    return {"ok": True, "id": item.id}


@app.post("/api/projects/{item_id}/restore")
def restore_project(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))]):
    item = db.get(Project, item_id)
    if not item:
        raise HTTPException(404, "Project not found")
    if item.name.startswith(HISTORICAL_PROJECT_PREFIX):
        raise HTTPException(400, "Imported historical projects cannot be restored")
    item.is_archived = False
    item.archived_at = None
    db.commit()
    return {"ok": True, "id": item.id}


@app.post("/api/shipments", response_model=ShipmentOut)
def create_shipment(payload: ShipmentBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    data = payload.model_dump(exclude={"items"})
    campaign = db.get(Campaign, payload.campaign_id)
    if not campaign:
        raise HTTPException(404, "Collaboration not found")
    apply_shipping_address_snapshot(db, campaign.media_id, data)
    item = Shipment(**data)
    db.add(item)
    db.flush()
    for row in payload.items:
        db.add(ShipmentItem(shipment_id=item.id, **row.model_dump()))
    db.commit()
    db.refresh(item)
    return item


@app.post("/api/projects/{project_id}/shipments", response_model=ShipmentOut)
def create_project_shipment(project_id: int, payload: ProjectShipmentBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    project = db.get(Project, project_id)
    media = db.get(Media, payload.media_id)
    if not project or not media:
        raise HTTPException(404, "Project or media not found")
    if payload.campaign_id:
        campaign = db.get(Campaign, payload.campaign_id)
        if not campaign or campaign.project_id != project_id or campaign.media_id != payload.media_id:
            raise HTTPException(400, "Selected collaboration does not belong to this project and media")
    else:
        campaign = db.query(Campaign).filter(Campaign.project_id == project_id, Campaign.media_id == payload.media_id, Campaign.execution_status.notin_(["已发布", "已结算", "已暂停", "已取消", "已暂停/取消"])).order_by(Campaign.updated_at.desc()).first()
    if not campaign:
        campaign = Campaign(project_id=project_id, media_id=payload.media_id, owner_id=payload.owner_id, execution_status="待发货", stage="Not Started", sample_status="Not Sent")
        db.add(campaign)
        db.flush()
        db.add(CampaignStageEvent(campaign_id=campaign.id, user_id=user.id, from_status=None, to_status="待发货", action="shipment_sync", reason="登记项目寄样时创建合作"))
    elif campaign.execution_status not in {"已发布", "已结算", "已暂停", "已取消", "已暂停/取消"} and payload.status in SHIPMENT_STATUS_PROGRESS:
        current_progress = SHIPMENT_STATUS_PROGRESS.get(campaign.execution_status, 0)
        next_progress = SHIPMENT_STATUS_PROGRESS.get(payload.status, current_progress)
        if next_progress > current_progress:
            set_campaign_status(db, campaign, user, payload.status, "shipment_sync", "登记物流信息自动同步")
    data = payload.model_dump(exclude={"items", "media_id", "campaign_id", "owner_id"})
    apply_shipping_address_snapshot(db, payload.media_id, data)
    shipment = Shipment(campaign_id=campaign.id, **data)
    db.add(shipment)
    db.flush()
    for row in payload.items:
        item_data = row.model_dump()
        if item_data.get("product_id"):
            product = db.get(Product, item_data["product_id"])
            if not product:
                raise HTTPException(400, "Product not found")
            item_data["product_name"] = product.model
            ensure_project_link(db, project_id, product.id)
        elif item_data.get("product_name"):
            product = find_or_create_product(db, item_data["product_name"], "项目寄样时创建")
            item_data["product_id"] = product.id
            item_data["product_name"] = product.model
            ensure_project_link(db, project_id, product.id)
        db.add(ShipmentItem(shipment_id=shipment.id, **item_data))
    db.commit()
    db.refresh(shipment)
    return shipment


@app.post("/api/cost-items", response_model=CostItemOut)
def create_cost_item(payload: CostItemBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    item = CostItem(**payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@app.post("/api/activities", response_model=ActivityOut)
def create_activity(payload: ActivityBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    item = Activity(**payload.model_dump(), user_id=user.id)
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@app.get("/api/campaigns", response_model=dict)
def list_campaigns(
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(current_user)],
    product_id: int | None = None,
    media_id: int | None = None,
    country: str | None = None,
    stage: str | None = None,
    sample_status: str | None = None,
    owner_id: int | None = None,
    history_only: bool = False,
    page: int = 1,
    page_size: int = 20,
):
    query = db.query(Campaign).options(joinedload(Campaign.project), joinedload(Campaign.product), joinedload(Campaign.media), joinedload(Campaign.owner), joinedload(Campaign.shipments), joinedload(Campaign.deliverables), joinedload(Campaign.cost_items)).outerjoin(Project)
    if history_only:
        query = query.filter(or_(Campaign.is_historical.is_(True), Project.is_archived.is_(True), Project.name.like(f"{HISTORICAL_PROJECT_PREFIX}%")))
    else:
        query = query.filter(Campaign.is_historical.is_(False), or_(Campaign.project_id.is_(None), and_(Project.is_archived.is_(False), ~Project.name.like(f"{HISTORICAL_PROJECT_PREFIX}%"))))
    if product_id:
        query = query.filter(Campaign.product_id == product_id)
    if media_id:
        query = query.filter(Campaign.media_id == media_id)
    if country:
        query = query.join(Media).filter(Media.country == country)
    if stage:
        query = query.filter(Campaign.stage == stage)
    if sample_status:
        query = query.filter(Campaign.sample_status == sample_status)
    if owner_id:
        query = query.filter(Campaign.owner_id == owner_id)
    ordered = query.order_by(Campaign.updated_at.desc())
    total = ordered.count()
    rows = ordered.offset((page - 1) * page_size).limit(page_size).all()
    items = []
    for row in rows:
        encoded = jsonable_encoder(row)
        encoded.update(collaboration_workflow_health(row))
        encoded.update(collaboration_advance_state(row))
        encoded.update(collaboration_stage_metadata(row))
        items.append(encoded)
    return {"items": items, "total": total}


@app.get("/api/collaborations/{item_id:int}")
def collaboration_detail(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(current_user)]):
    item = db.query(Campaign).options(joinedload(Campaign.project), joinedload(Campaign.product), joinedload(Campaign.media).joinedload(Media.contacts), joinedload(Campaign.owner), joinedload(Campaign.shipments).joinedload(Shipment.items), joinedload(Campaign.deliverables).joinedload(Deliverable.performance_snapshots), joinedload(Campaign.cost_items), joinedload(Campaign.activities).joinedload(Activity.user), joinedload(Campaign.stage_events).joinedload(CampaignStageEvent.user)).filter(Campaign.id == item_id).first()
    if not item:
        raise HTTPException(404, "Collaboration not found")
    result = jsonable_encoder(item)
    result["deliverables"] = [deliverable_detail_payload(row) for row in item.deliverables]
    result.update(collaboration_workflow_health(item))
    result.update(collaboration_advance_state(item))
    result.update(collaboration_stage_metadata(item))
    return result


@app.post("/api/collaborations/{item_id:int}/advance")
def advance_collaboration(
    item_id: int,
    payload: CollaborationAdvanceIn,
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(editable_user)],
):
    item = db.query(Campaign).options(
        joinedload(Campaign.project),
        joinedload(Campaign.media),
        joinedload(Campaign.owner),
        joinedload(Campaign.shipments),
        joinedload(Campaign.deliverables),
        joinedload(Campaign.cost_items),
    ).filter(Campaign.id == item_id).first()
    if not item:
        raise HTTPException(404, "Collaboration not found")
    before_status = item.execution_status
    advance = collaboration_advance_state(item)
    if not advance["next_status"]:
        raise HTTPException(409, "当前阶段没有可推进的下一阶段")
    if payload.target_status != advance["next_status"]:
        raise HTTPException(409, f"只能从“{before_status}”推进到相邻阶段“{advance['next_status']}”")

    shipment = item.shipments[0] if item.shipments else None
    if payload.tracking_number or payload.carrier or payload.delivered_at:
        if shipment is None:
            shipment = Shipment(campaign_id=item.id, status=before_status)
            db.add(shipment)
            item.shipments.append(shipment)
        if payload.tracking_number:
            shipment.tracking_number = payload.tracking_number.strip()
        if payload.carrier:
            shipment.carrier = payload.carrier.strip()
        if payload.delivered_at:
            shipment.delivered_at = payload.delivered_at

    deliverable = item.deliverables[0] if item.deliverables else None
    if payload.content_title or payload.publication_url or payload.published_at:
        if deliverable is None:
            deliverable = Deliverable(campaign_id=item.id, deliverable_type="Other")
            db.add(deliverable)
            item.deliverables.append(deliverable)
        if payload.content_title:
            deliverable.title = payload.content_title.strip()
        if payload.publication_url:
            deliverable.url = payload.publication_url.strip()
        if payload.published_at:
            deliverable.published_at = payload.published_at

    if payload.target_status == "已结算" and not item.cost_items and payload.no_payment_required:
        cost = CostItem(campaign_id=item.id, cost_type="无需付款", planned_amount=0, actual_amount=0, payment_status="无需付款")
        db.add(cost)
        item.cost_items.append(cost)

    db.flush()
    checked = collaboration_advance_state(item)
    if checked["advance_blockers"]:
        db.rollback()
        raise HTTPException(409, {"message": "尚不能推进", "blockers": checked["advance_blockers"], "requirements": checked["advance_requirements"]})

    set_campaign_status(db, item, user, payload.target_status, "advance")
    if payload.follow_up_date is not None:
        item.follow_up_date = payload.follow_up_date
    if payload.target_status == "已发布":
        item.stage = "Published"
        item.actual_publish_date = payload.published_at or (deliverable.published_at if deliverable else None)
    if shipment and payload.target_status in SHIPMENT_STATUS_PROGRESS:
        shipment.status = payload.target_status
    add_audit_log(db, user, "advance", "collaboration", item.id, before={"execution_status": before_status}, after=payload.model_dump(mode="json", exclude_none=True))
    db.commit()
    refreshed = db.query(Campaign).options(
        joinedload(Campaign.project), joinedload(Campaign.media), joinedload(Campaign.owner),
        joinedload(Campaign.shipments), joinedload(Campaign.deliverables), joinedload(Campaign.cost_items), joinedload(Campaign.activities),
    ).filter(Campaign.id == item.id).first()
    result = jsonable_encoder(refreshed)
    result.update(collaboration_workflow_health(refreshed))
    result.update(collaboration_advance_state(refreshed))
    result.update(collaboration_stage_metadata(refreshed))
    return result


@app.post("/api/collaborations/{item_id:int}/status-action")
def collaboration_status_action(
    item_id: int,
    payload: CollaborationStatusActionIn,
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(editable_user)],
):
    item = db.query(Campaign).options(joinedload(Campaign.stage_events)).filter(Campaign.id == item_id).first()
    if not item:
        raise HTTPException(404, "Collaboration not found")
    reason = (payload.reason or "").strip() or None
    action = payload.action
    target: str | None = None
    if action == "pause":
        if item.execution_status in FOLLOW_UP_CLOSED_STATUSES:
            raise HTTPException(409, "当前阶段不能暂停")
        target = "已暂停"
    elif action == "cancel":
        if item.execution_status in {"已结算", "已取消"}:
            raise HTTPException(409, "当前阶段不能取消")
        if not reason or len(reason) < 2:
            raise HTTPException(400, "取消合作时请填写具体原因")
        target = "已取消"
    elif action == "resume":
        if item.execution_status not in {"已暂停", "已暂停/取消"}:
            raise HTTPException(409, "只有已暂停合作可以恢复")
        paused_event = next((event for event in sorted(item.stage_events, key=lambda row: row.created_at, reverse=True) if event.to_status in {"已暂停", "已暂停/取消"} and event.from_status in EXECUTION_SEQUENCE), None)
        target = paused_event.from_status if paused_event else "待确认"
    elif action == "rollback":
        if not reason or len(reason) < 2:
            raise HTTPException(400, "回退阶段时请填写具体原因")
        if item.execution_status not in EXECUTION_SEQUENCE or EXECUTION_SEQUENCE.index(item.execution_status) == 0:
            raise HTTPException(409, "当前阶段不能回退")
        target = EXECUTION_SEQUENCE[EXECUTION_SEQUENCE.index(item.execution_status) - 1]
    elif action == "override":
        if user.role != "Admin":
            raise HTTPException(403, "仅管理员可以指定阶段")
        if payload.target_status not in EXECUTION_STATUSES:
            raise HTTPException(400, "请选择有效目标阶段")
        target = payload.target_status
    else:
        raise HTTPException(400, "不支持的状态操作")
    before = item.execution_status
    set_campaign_status(db, item, user, target, action, reason)
    add_audit_log(db, user, f"status_{action}", "collaboration", item.id, before={"execution_status": before}, after={"execution_status": target}, reason=reason)
    db.commit()
    refreshed = db.query(Campaign).options(joinedload(Campaign.project), joinedload(Campaign.media), joinedload(Campaign.owner), joinedload(Campaign.shipments), joinedload(Campaign.deliverables), joinedload(Campaign.cost_items), joinedload(Campaign.activities), joinedload(Campaign.stage_events)).filter(Campaign.id == item.id).first()
    result = jsonable_encoder(refreshed)
    result.update(collaboration_workflow_health(refreshed))
    result.update(collaboration_advance_state(refreshed))
    result.update(collaboration_stage_metadata(refreshed))
    return result


@app.post("/api/collaborations/{item_id:int}/undo-advance")
def undo_collaboration_advance(
    item_id: int,
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(editable_user)],
):
    item = db.query(Campaign).options(joinedload(Campaign.stage_events)).filter(Campaign.id == item_id).first()
    if not item:
        raise HTTPException(404, "Collaboration not found")
    latest = max(item.stage_events, key=lambda row: row.created_at) if item.stage_events else None
    if not latest or latest.action != "advance" or latest.to_status != item.execution_status:
        raise HTTPException(409, "当前没有可撤销的推进")
    if latest.user_id != user.id:
        raise HTTPException(403, "只能撤销自己刚刚执行的推进")
    if datetime.utcnow() - latest.created_at > timedelta(seconds=30):
        raise HTTPException(409, "撤销时间已超过 30 秒；如需调整请使用回退一步")
    target = latest.from_status
    if target not in EXECUTION_STATUSES:
        raise HTTPException(409, "原阶段不可恢复")
    before = item.execution_status
    set_campaign_status(db, item, user, target, "undo", "撤销刚刚的阶段推进")
    add_audit_log(db, user, "undo_advance", "collaboration", item.id, before={"execution_status": before}, after={"execution_status": target})
    db.commit()
    refreshed = db.query(Campaign).options(joinedload(Campaign.project), joinedload(Campaign.media), joinedload(Campaign.owner), joinedload(Campaign.shipments), joinedload(Campaign.deliverables), joinedload(Campaign.cost_items), joinedload(Campaign.activities), joinedload(Campaign.stage_events)).filter(Campaign.id == item.id).first()
    result = jsonable_encoder(refreshed)
    result.update(collaboration_workflow_health(refreshed))
    result.update(collaboration_advance_state(refreshed))
    result.update(collaboration_stage_metadata(refreshed))
    return result


@app.patch("/api/collaborations/{item_id:int}")
def patch_collaboration(
    item_id: int,
    payload: CollaborationPatch,
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(editable_user)],
    change_reason: Annotated[str | None, Header(alias="X-Change-Reason")] = None,
):
    item = db.query(Campaign).options(joinedload(Campaign.shipments)).filter(Campaign.id == item_id).first()
    if not item:
        raise HTTPException(404, "Collaboration not found")
    data = payload.model_dump(exclude_unset=True)
    if "execution_status" in data and data["execution_status"] != item.execution_status:
        raise HTTPException(409, "执行状态不能直接修改，请使用推进到下一阶段或状态操作")
    before = {key: jsonable_encoder(getattr(item, key, None)) for key in data}
    for key in ["project_id", "media_id", "owner_id", "collaboration_type", "expected_publish_date", "notes", "next_action", "follow_up_date", "follow_up_priority", "follow_up_done"]:
        if key in data:
            setattr(item, key, data[key])
    if data.get("follow_up_done") is True:
        db.add(Activity(campaign_id=item.id, user_id=user.id, activity_type="待办完成", content=f"已完成待办：{item.next_action or '未填写下一步动作'}"))
    if "tracking_number" in data or "oa_pi_number" in data:
        shipment = item.shipments[0] if item.shipments else Shipment(campaign_id=item.id, status=item.execution_status)
        if not item.shipments:
            db.add(shipment)
        if "tracking_number" in data:
            shipment.tracking_number = data["tracking_number"]
        if "oa_pi_number" in data:
            shipment.oa_pi_number = data["oa_pi_number"]
    add_audit_log(db, user, "update", "collaboration", item.id, before=before, after=data, reason=change_reason)
    db.commit()
    refreshed = db.query(Campaign).options(joinedload(Campaign.project), joinedload(Campaign.media), joinedload(Campaign.owner), joinedload(Campaign.shipments)).filter(Campaign.id == item_id).first()
    result = {
        "id": refreshed.id,
        "project_id": refreshed.project_id,
        "project_name": refreshed.project.name if refreshed.project else "未归属项目",
        "media_id": refreshed.media_id,
        "media_name": refreshed.media.name if refreshed.media else None,
        "owner_id": refreshed.owner_id,
        "owner": refreshed.owner.name if refreshed.owner else None,
        "execution_status": refreshed.execution_status,
        "expected_publish_date": refreshed.expected_publish_date,
        "next_action": refreshed.next_action,
        "follow_up_date": refreshed.follow_up_date,
        "follow_up_priority": refreshed.follow_up_priority,
        "follow_up_done": refreshed.follow_up_done,
        "oa_pi_number": refreshed.shipments[0].oa_pi_number if refreshed.shipments else None,
        "tracking_number": refreshed.shipments[0].tracking_number if refreshed.shipments else None,
        "notes": refreshed.notes,
    }
    result.update(collaboration_workflow_health(refreshed))
    result.update(collaboration_advance_state(refreshed))
    result.update(collaboration_stage_metadata(refreshed))
    return result


@app.patch("/api/collaborations/bulk")
def bulk_patch_collaborations(payload: CollaborationBulkPatch, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    ids = list(dict.fromkeys(payload.ids))
    if not ids:
        raise HTTPException(400, "请选择至少一条执行单")
    data = payload.model_dump(exclude_unset=True, exclude={"ids", "preview"})
    if not data:
        raise HTTPException(400, "请选择要修改的字段")
    if "execution_status" in data:
        raise HTTPException(409, "批量操作不能修改执行状态，请逐条推进")
    items = db.query(Campaign).filter(Campaign.id.in_(ids)).all()
    preview_rows = []
    for item in items:
        before = {key: getattr(item, key) for key in data}
        after = {**before, **data}
        preview_rows.append({"id": item.id, "media": item.media.name if item.media else None, "before": before, "after": after})
    if payload.preview:
        return {"preview": True, "updated": 0, "matched": len(items), "items": jsonable_encoder(preview_rows)}
    for item in items:
        for key, value in data.items():
            setattr(item, key, value)
        db.add(Activity(campaign_id=item.id, user_id=user.id, activity_type="批量更新", content="已通过工作台批量更新"))
    db.commit()
    return {"preview": False, "updated": len(items), "items": jsonable_encoder(preview_rows)}


@app.post("/api/campaigns", response_model=CampaignOut)
def create_campaign(payload: CampaignBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    validate_campaign(payload)
    if payload.execution_status != "待确认":
        raise HTTPException(409, "新合作必须从“待确认”开始，历史数据请通过导入流程写入")
    data = payload.model_dump()
    if "next_action" not in payload.model_fields_set:
        data["next_action"] = NEXT_ACTION_BY_STATUS.get(data.get("execution_status"))
    item = Campaign(**data)
    db.add(item)
    db.flush()
    db.add(CampaignStageEvent(campaign_id=item.id, user_id=user.id, from_status=None, to_status=item.execution_status, action="create", reason="新建合作"))
    db.add(Activity(campaign_id=item.id, user_id=user.id, activity_type="创建执行单", content=f"创建合作执行单，当前阶段：{item.execution_status}"))
    db.commit()
    db.refresh(item)
    return item


@app.put("/api/campaigns/{item_id}", response_model=CampaignOut)
def update_campaign(item_id: int, payload: CampaignBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    validate_campaign(payload)
    item = db.get(Campaign, item_id)
    if not item:
        raise HTTPException(404, "Campaign not found")
    if payload.execution_status != item.execution_status:
        raise HTTPException(409, "执行状态不能直接修改，请使用推进到下一阶段或状态操作")
    for key, value in payload.model_dump().items():
        setattr(item, key, value)
    db.commit()
    db.refresh(item)
    return item


@app.delete("/api/campaigns/{item_id}")
def delete_campaign(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))]):
    item = db.query(Campaign).options(joinedload(Campaign.shipments), joinedload(Campaign.cost_items), joinedload(Campaign.deliverables), joinedload(Campaign.activities)).filter(Campaign.id == item_id).first()
    if not item:
        raise HTTPException(404, "Campaign not found")
    counts = {"deleted_shipments": len(item.shipments), "deleted_cost_items": len(item.cost_items), "deleted_deliverables": len(item.deliverables), "deleted_activities": len(item.activities)}
    db.delete(item)
    db.commit()
    return {"ok": True, **counts}


@app.post("/api/campaigns/{item_id}/archive")
def archive_campaign(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))]):
    item = db.get(Campaign, item_id)
    if not item:
        raise HTTPException(404, "Campaign not found")
    item.is_historical = True
    item.archived_at = datetime.now()
    db.commit()
    return {"ok": True, "id": item.id}


@app.post("/api/campaigns/{item_id}/restore")
def restore_campaign(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))]):
    item = db.query(Campaign).options(joinedload(Campaign.project)).filter(Campaign.id == item_id).first()
    if not item:
        raise HTTPException(404, "Campaign not found")
    if not item.archived_at:
        raise HTTPException(400, "Imported historical collaborations cannot be restored")
    if item.project and item.project.is_archived:
        raise HTTPException(400, "Restore the project before restoring this collaboration")
    item.is_historical = False
    item.archived_at = None
    db.commit()
    return {"ok": True, "id": item.id}


@app.get("/api/deliverables", response_model=dict)
def list_deliverables(
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(current_user)],
    product_id: int | None = None,
    media_id: int | None = None,
    deliverable_type: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    page: int = 1,
    page_size: int = 20,
):
    query = db.query(Deliverable).options(joinedload(Deliverable.campaign).joinedload(Campaign.product), joinedload(Deliverable.campaign).joinedload(Campaign.media))
    query = query.join(Campaign)
    if product_id:
        query = query.filter(Campaign.product_id == product_id)
    if media_id:
        query = query.filter(Campaign.media_id == media_id)
    if deliverable_type:
        query = query.filter(Deliverable.deliverable_type == deliverable_type)
    if date_from:
        query = query.filter(Deliverable.published_at >= date_from)
    if date_to:
        query = query.filter(Deliverable.published_at <= date_to)
    return list_payload(query.order_by(Deliverable.published_at.desc().nullslast()), page, page_size)


@app.get("/api/content-monitor/status")
def content_monitor_status(user: Annotated[User, Depends(current_user)]):
    return monitor_runtime_status()


@app.post("/api/content-monitor/run")
def run_content_monitor_now(db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))]):
    try:
        return run_content_monitor(db)
    except (YouTubeConfigurationError, YouTubeSourceError) as exc:
        raise HTTPException(502, str(exc)) from exc


@app.post("/api/deliverables", response_model=DeliverableOut)
def create_deliverable(payload: DeliverableBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    if payload.deliverable_type not in DELIVERABLE_TYPES:
        raise HTTPException(400, "Invalid deliverable type")
    item = Deliverable(**payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@app.put("/api/deliverables/{item_id}", response_model=DeliverableOut)
def update_deliverable(item_id: int, payload: DeliverableBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    item = db.get(Deliverable, item_id)
    if not item:
        raise HTTPException(404, "Deliverable not found")
    for key, value in payload.model_dump().items():
        setattr(item, key, value)
    db.commit()
    db.refresh(item)
    return item


@app.get("/api/contacts", response_model=dict)
def list_contacts(db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(current_user)], q: str | None = None, media_id: int | None = None, country: str | None = None, page: int = 1, page_size: int = 20):
    query = db.query(Contact).options(joinedload(Contact.media)).join(Media)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Contact.name.ilike(like), Contact.email.ilike(like), Contact.notes.ilike(like), Media.name.ilike(like)))
    if media_id:
        query = query.filter(Contact.media_id == media_id)
    if country:
        query = query.filter(Media.country == country)
    return list_payload(query.order_by(Contact.id.desc()), page, page_size)


def normalized_email(value: str | None) -> str | None:
    return (value or "").strip().casefold() or None


def normalized_phone(value: str | None) -> str | None:
    digits = re.sub(r"\D+", "", value or "")
    return digits or None


def normalize_contact_data(data: dict[str, Any]) -> dict[str, Any]:
    for field in ("email", "brief_email", "press_release_email"):
        data[field] = normalized_email(data.get(field))
    for field in ("phone", "whatsapp"):
        data[field] = (data.get(field) or "").strip() or None
    return data


def require_unique_contact(db: Session, data: dict[str, Any], exclude_id: int | None = None) -> None:
    emails = {normalized_email(data.get(field)) for field in ("email", "brief_email", "press_release_email")} - {None}
    phones = {normalized_phone(data.get(field)) for field in ("phone", "whatsapp")} - {None}
    matches: list[dict[str, Any]] = []
    for item in db.query(Contact).all():
        if item.id == exclude_id:
            continue
        item_emails = {normalized_email(getattr(item, field)) for field in ("email", "brief_email", "press_release_email")} - {None}
        item_phones = {normalized_phone(getattr(item, field)) for field in ("phone", "whatsapp")} - {None}
        reasons = [*(f"邮箱：{value}" for value in sorted(emails & item_emails)), *(f"电话：{value}" for value in sorted(phones & item_phones))]
        if reasons:
            matches.append({"id": item.id, "media_id": item.media_id, "name": item.name, "reasons": reasons})
    if matches:
        raise HTTPException(409, detail={"message": "联系人邮箱或电话已存在，请使用现有联系人或先合并", "matches": matches})


@app.post("/api/contacts", response_model=ContactOut)
def create_contact(payload: ContactBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    if not db.get(Media, payload.media_id):
        raise HTTPException(404, "Media not found")
    data = normalize_contact_data(payload.model_dump())
    data["data_capture_method"] = data.get("data_capture_method") or "manual"
    data["data_source"] = data.get("data_source") or "CRM 人工录入"
    require_unique_contact(db, data)
    if payload.is_primary:
        db.query(Contact).filter(Contact.media_id == payload.media_id).update({Contact.is_primary: False}, synchronize_session=False)
    item = Contact(**data)
    db.add(item)
    db.flush()
    add_audit_log(db, user, "create", "contact", item.id, after=data, reason="创建联系人")
    db.commit()
    db.refresh(item)
    return item


@app.put("/api/contacts/{item_id}", response_model=ContactOut)
def update_contact(item_id: int, payload: ContactBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    item = db.get(Contact, item_id)
    if not item:
        raise HTTPException(404, "Contact not found")
    if not db.get(Media, payload.media_id):
        raise HTTPException(404, "Media not found")
    before = ContactOut.model_validate(item).model_dump(mode="json")
    data = normalize_contact_data(payload.model_dump())
    require_unique_contact(db, data, item_id)
    if payload.is_primary:
        db.query(Contact).filter(Contact.media_id == payload.media_id, Contact.id != item_id).update({Contact.is_primary: False}, synchronize_session=False)
    for key, value in data.items():
        setattr(item, key, value)
    add_audit_log(db, user, "update", "contact", item.id, before=before, after=data, reason="更新联系人")
    db.commit()
    db.refresh(item)
    return item


@app.delete("/api/contacts/{item_id}")
def delete_contact(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    item = db.get(Contact, item_id)
    if not item:
        raise HTTPException(404, "Contact not found")
    before = ContactOut.model_validate(item).model_dump(mode="json")
    db.query(ShippingAddress).filter(ShippingAddress.contact_id == item_id).update({ShippingAddress.contact_id: None}, synchronize_session=False)
    add_audit_log(db, user, "delete", "contact", item.id, before=before, reason="删除联系人")
    db.delete(item)
    db.commit()
    return {"ok": True}


@app.get("/api/media/{media_id}/shipping-addresses", response_model=list[ShippingAddressOut])
def list_shipping_addresses(media_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(current_user)]):
    if not db.get(Media, media_id):
        raise HTTPException(404, "Media not found")
    return db.query(ShippingAddress).filter(ShippingAddress.media_id == media_id).order_by(ShippingAddress.is_default.desc(), ShippingAddress.updated_at.desc()).all()


@app.post("/api/shipping-addresses", response_model=ShippingAddressOut)
def create_shipping_address(payload: ShippingAddressBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    validate_shipping_address_links(db, payload.media_id, payload.contact_id)
    item = ShippingAddress(**payload.model_dump())
    db.add(item)
    db.flush()
    has_default = db.query(ShippingAddress.id).filter(ShippingAddress.media_id == payload.media_id, ShippingAddress.is_default.is_(True), ShippingAddress.id != item.id).first()
    if payload.is_default or not has_default:
        set_default_shipping_address(db, item)
    db.commit()
    db.refresh(item)
    return item


@app.put("/api/shipping-addresses/{item_id}", response_model=ShippingAddressOut)
def update_shipping_address(item_id: int, payload: ShippingAddressBase, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    item = db.get(ShippingAddress, item_id)
    if not item:
        raise HTTPException(404, "Shipping address not found")
    validate_shipping_address_links(db, payload.media_id, payload.contact_id)
    old_media_id = item.media_id
    for key, value in payload.model_dump().items():
        setattr(item, key, value)
    if payload.is_default:
        set_default_shipping_address(db, item)
    db.flush()
    if old_media_id != item.media_id and not db.query(ShippingAddress.id).filter(ShippingAddress.media_id == old_media_id, ShippingAddress.is_default.is_(True)).first():
        fallback = db.query(ShippingAddress).filter(ShippingAddress.media_id == old_media_id).order_by(ShippingAddress.updated_at.desc()).first()
        if fallback:
            fallback.is_default = True
    db.commit()
    db.refresh(item)
    return item


@app.post("/api/shipping-addresses/{item_id}/default", response_model=ShippingAddressOut)
def make_default_shipping_address(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    item = db.get(ShippingAddress, item_id)
    if not item:
        raise HTTPException(404, "Shipping address not found")
    set_default_shipping_address(db, item)
    db.commit()
    db.refresh(item)
    return item


@app.delete("/api/shipping-addresses/{item_id}")
def delete_shipping_address(item_id: int, db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    item = db.get(ShippingAddress, item_id)
    if not item:
        raise HTTPException(404, "Shipping address not found")
    media_id, was_default = item.media_id, item.is_default
    db.query(Shipment).filter(Shipment.shipping_address_id == item_id).update({Shipment.shipping_address_id: None}, synchronize_session=False)
    db.delete(item)
    db.flush()
    if was_default:
        fallback = db.query(ShippingAddress).filter(ShippingAddress.media_id == media_id).order_by(ShippingAddress.updated_at.desc()).first()
        if fallback:
            fallback.is_default = True
    db.commit()
    return {"ok": True}


def parse_address_candidate(raw: str, country: str | None = None) -> dict[str, Any]:
    email = re.search(r"[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}", raw)
    phone = re.search(r"(?:\+?\d[\d\s().-]{6,}\d)", raw)
    postal = re.search(r"(?i)(?:zip|postal(?:\s*code)?|邮编)\s*[:：]?\s*([A-Z0-9 -]{3,12})", raw)
    recipient = re.search(r"(?i)(?:recipient|receiver|contact|收件人|姓名)\s*[:：]\s*([^\n,;]+)", raw)
    return {
        "recipient_name": recipient.group(1).strip() if recipient else None,
        "phone": phone.group(0).strip() if phone else None,
        "email": email.group(0) if email else None,
        "address_text": raw.strip(),
        "postal_code": postal.group(1).strip() if postal else None,
        "country": country,
        "source_text": raw.strip(),
        "is_confirmed": True,
    }


@app.get("/api/address-import/candidates")
def address_import_candidates(db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(editable_user)]):
    source = ROOT / "outputs" / "cleaning" / "铭瑄红人记者库_cleaned.xlsx.inspect.ndjson"
    if not source.exists():
        return {"source_available": False, "items": []}
    keywords = re.compile(r"(?i)address|street|road|avenue|building|floor|postal|zip|地址|邮编|收件|省|市|区")
    source_rows: list[dict[str, Any]] = []
    with source.open("r", encoding="utf-8") as handle:
        for line in handle:
            try:
                block = json.loads(line)
            except json.JSONDecodeError:
                continue
            if block.get("kind") != "table" or block.get("sheet") != "Raw 原始数据":
                continue
            values = block.get("values") or []
            if not values:
                continue
            headers = [str(value or "") for value in values[0]]
            for row in values[1:]:
                source_rows.append({headers[index]: row[index] if index < len(row) else None for index in range(len(headers))})
    items: list[dict[str, Any]] = []
    for row_number, flat in enumerate(source_rows, 2):
        raw = str(flat.get("联系方式") or "").strip()
        if not raw or not keywords.search(raw):
            continue
        media_name = str(flat.get("名字") or flat.get("名称") or "").strip()
        country = str(flat.get("国家") or "").strip() or None
        contact_name = str(flat.get("联系人&职位") or flat.get("联系人") or "").strip()
        media = db.query(Media).filter(func.lower(Media.name) == media_name.lower()).first() if media_name else None
        contact = None
        if media and contact_name:
            contact = db.query(Contact).filter(Contact.media_id == media.id, Contact.name.ilike(f"%{contact_name.split('/')[0].strip()}%")).first()
        imported = db.query(ShippingAddress.id).filter(ShippingAddress.source_text == raw).first() is not None
        items.append({"id": f"row-{row_number}", "media_name": media_name, "contact_name": contact_name, "raw_text": raw, "media_id": media.id if media else None, "contact_id": contact.id if contact else None, "imported": imported, "parsed": parse_address_candidate(raw, country)})
    return {"source_available": True, "items": items}


@app.get("/api/dashboard")
def dashboard(
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(current_user)],
    product_model: str | None = None,
    country: str | None = None,
    platform_type: str | None = None,
    stage: str | None = None,
    sample_status: str | None = None,
    owner_id: int | None = None,
    has_deliverable: bool | None = None,
):
    base = db.query(Campaign).options(joinedload(Campaign.product), joinedload(Campaign.media), joinedload(Campaign.owner), joinedload(Campaign.deliverables))
    if product_model:
        like = f"%{product_model}%"
        base = base.join(Product, isouter=True).filter(or_(Product.model.ilike(like), Product.full_name.ilike(like), Product.aliases.ilike(like)))
    if country or platform_type:
        base = base.join(Media)
    if country:
        base = base.filter(Media.country == country)
    if platform_type:
        base = base.filter(Media.platform_type == platform_type)
    if stage:
        base = base.filter(Campaign.stage == stage)
    if sample_status:
        base = base.filter(Campaign.sample_status == sample_status)
    if owner_id:
        base = base.filter(Campaign.owner_id == owner_id)
    if has_deliverable is True:
        base = base.join(Deliverable)
    elif has_deliverable is False:
        base = base.outerjoin(Deliverable).filter(Deliverable.id.is_(None))
    campaigns = base.order_by(Campaign.updated_at.desc()).limit(200).all()
    now_count = lambda condition: db.query(func.count(Campaign.id)).filter(condition).scalar() or 0
    kpis = {
        "media_total": db.query(func.count(Media.id)).scalar() or 0,
        "product_total": db.query(func.count(Product.id)).scalar() or 0,
        "campaign_total": db.query(func.count(Campaign.id)).scalar() or 0,
        "contacted_total": now_count(Campaign.stage.in_(["Contacted", "Waiting Reply", "Quoting", "Brief Sent", "Sample Sent", "In Production", "Published"])),
        "brief_sent_total": now_count(Campaign.brief_sent.is_(True)),
        "sample_sent_total": now_count(Campaign.sample_status.in_(["Shipped", "In Transit", "Customs Clearance", "Delivered"])),
        "in_production_total": now_count(Campaign.stage == "In Production"),
        "published_total": now_count(Campaign.stage == "Published"),
        "overdue_total": db.query(func.count(Campaign.id)).filter(and_(Campaign.expected_publish_date < func.date("now"), Campaign.actual_publish_date.is_(None))).scalar() or 0,
    }
    rows = []
    for campaign in campaigns:
        first_deliverable = campaign.deliverables[0] if campaign.deliverables else None
        rows.append(
            {
                "id": campaign.id,
                "product_model": campaign.product.model if campaign.product else None,
                "media_name": campaign.media.name if campaign.media else None,
                "country": campaign.media.country if campaign.media else None,
                "platform_type": campaign.media.platform_type if campaign.media else None,
                "owner": campaign.owner.name if campaign.owner else None,
                "stage": campaign.stage,
                "sample_status": campaign.sample_status,
                "brief_sent": campaign.brief_sent,
                "expected_publish_date": campaign.expected_publish_date,
                "actual_publish_date": campaign.actual_publish_date,
                "deliverable_url": first_deliverable.url if first_deliverable else None,
                "views": first_deliverable.views if first_deliverable else None,
            }
        )
    return {"kpis": kpis, "items": rows}


@app.get("/api/workbench")
def workbench(
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(current_user)],
    project_id: int | None = None,
    owner_id: int | None = None,
    execution_status: str | None = None,
    country: str | None = None,
    platform_type: str | None = None,
    payment_pending: bool = False,
    queue: str = "today",
):
    query = db.query(Campaign).options(joinedload(Campaign.project), joinedload(Campaign.media), joinedload(Campaign.owner), joinedload(Campaign.shipments), joinedload(Campaign.deliverables), joinedload(Campaign.cost_items)).outerjoin(Project).filter(Campaign.is_historical.is_(False), or_(Campaign.project_id.is_(None), and_(Project.is_archived.is_(False), ~Project.name.like(f"{HISTORICAL_PROJECT_PREFIX}%"))))
    if project_id:
        query = query.filter(Campaign.project_id == project_id)
    if owner_id:
        query = query.filter(Campaign.owner_id == owner_id)
    if execution_status:
        query = query.filter(Campaign.execution_status == execution_status)
    if payment_pending:
        query = query.filter(Campaign.cost_items.any(CostItem.payment_status.in_(["未付款", "部分付款"])))
    if country or platform_type:
        query = query.join(Media)
    if country:
        query = query.filter(Media.country == country)
    if platform_type:
        query = query.filter(Media.platform_type == platform_type)
    today_date = date.today()
    if queue == "overdue":
        query = query.filter(Campaign.follow_up_done.is_(False), Campaign.follow_up_date < today_date)
    elif queue == "today":
        query = query.filter(Campaign.follow_up_done.is_(False), Campaign.follow_up_date == today_date)
    elif queue == "upcoming":
        query = query.filter(Campaign.follow_up_done.is_(False), Campaign.follow_up_date > today_date, Campaign.follow_up_date <= today_date + timedelta(days=7))
    elif queue != "all":
        raise HTTPException(400, "Invalid queue")
    items = query.order_by(Campaign.follow_up_date.asc().nullslast(), Campaign.updated_at.desc()).limit(300).all()
    today = func.date("now")
    visible_campaigns = db.query(Campaign.id).outerjoin(Project).filter(Campaign.is_historical.is_(False), or_(Campaign.project_id.is_(None), and_(Project.is_archived.is_(False), ~Project.name.like(f"{HISTORICAL_PROJECT_PREFIX}%"))))
    overdue = visible_campaigns.filter(Campaign.expected_publish_date < today, Campaign.actual_publish_date.is_(None), Campaign.execution_status.notin_(["已发布", "已结算", "已暂停", "已取消", "已暂停/取消"])).count()
    all_costs = db.query(CostItem).join(Campaign).outerjoin(Project).filter(Campaign.is_historical.is_(False), or_(Campaign.project_id.is_(None), and_(Project.is_archived.is_(False), ~Project.name.like(f"{HISTORICAL_PROJECT_PREFIX}%")))).all()
    kpi_campaigns = db.query(Campaign).options(joinedload(Campaign.cost_items)).outerjoin(Project).filter(Campaign.is_historical.is_(False), or_(Campaign.project_id.is_(None), and_(Project.is_archived.is_(False), ~Project.name.like(f"{HISTORICAL_PROJECT_PREFIX}%")))).all()
    rows = []
    for item in items:
        actual = sum(cost.actual_amount or 0 for cost in item.cost_items)
        planned = sum(cost.planned_amount or 0 for cost in item.cost_items)
        pending_payment = any(cost.payment_status in ["未付款", "部分付款"] for cost in item.cost_items)
        rows.append({
            "id": item.id,
            "project_id": item.project_id,
            "project_name": item.project.name if item.project else "未归属项目",
            "media_name": item.media.name if item.media else None,
            "country": item.media.country if item.media else None,
            "platform_type": item.media.platform_type if item.media else None,
            "owner": item.owner.name if item.owner else None,
            "execution_status": item.execution_status,
            "next_action": item.next_action,
            "follow_up_date": item.follow_up_date,
            "follow_up_priority": item.follow_up_priority,
            "follow_up_done": item.follow_up_done,
            "expected_publish_date": item.expected_publish_date,
            "tracking_number": item.shipments[0].tracking_number if item.shipments else None,
            "actual_amount": actual,
            "planned_amount": planned,
            "pending_payment": pending_payment,
            "content_url": item.deliverables[0].url if item.deliverables else None,
            **collaboration_workflow_health(item),
            **collaboration_advance_state(item),
            **collaboration_stage_metadata(item),
        })
    return {
        "kpis": {
            "project_total": db.query(func.count(Project.id)).scalar() or 0,
            "collaboration_total": len(kpi_campaigns),
            "pending_shipping": sum(1 for item in kpi_campaigns if item.execution_status == "待发货"),
            "in_transit": sum(1 for item in kpi_campaigns if item.execution_status == "运输中"),
            "awaiting_content": sum(1 for item in kpi_campaigns if item.execution_status == "已签收待产出"),
            "published": sum(1 for item in kpi_campaigns if item.execution_status == "已发布"),
            "overdue_content": overdue,
            "actual_amount": sum(cost.actual_amount or 0 for cost in all_costs),
            "pending_payment": sum(1 for item in kpi_campaigns if any(cost.payment_status in ["未付款", "部分付款"] for cost in item.cost_items)),
            "overdue_tasks": sum(1 for item in kpi_campaigns if not item.follow_up_done and item.follow_up_date and item.follow_up_date < today_date),
            "today_tasks": sum(1 for item in kpi_campaigns if not item.follow_up_done and item.follow_up_date == today_date),
            "upcoming_tasks": sum(1 for item in kpi_campaigns if not item.follow_up_done and item.follow_up_date and today_date < item.follow_up_date <= today_date + timedelta(days=7)),
        },
        "items": rows,
    }


@app.get("/api/import-batches")
def list_import_batches(
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(require_roles("Admin"))],
    limit: int = Query(30, ge=1, le=100),
):
    rows = db.query(ImportBatch).order_by(ImportBatch.created_at.desc()).limit(limit).all()
    return {"items": [{
        "id": row.id, "import_type": row.import_type, "filename": row.filename,
        "source_hash": row.source_hash, "status": row.status,
        "summary": json.loads(row.summary_json) if row.summary_json else {},
        "created_at": row.created_at, "undone_at": row.undone_at,
        "user": row.user.name if row.user else None,
    } for row in rows]}


@app.post("/api/import-batches/{batch_id}/undo")
def undo_import_batch(
    batch_id: int,
    db: Annotated[Session, Depends(get_db)],
    user: Annotated[User, Depends(require_roles("Admin"))],
    change_reason: Annotated[str | None, Header(alias="X-Change-Reason")] = None,
):
    batch = db.get(ImportBatch, batch_id)
    if not batch:
        raise HTTPException(404, "Import batch not found")
    if batch.undone_at:
        raise HTTPException(409, "该导入批次已撤销")
    model_map = {
        "media": Media, "contact": Contact, "campaign": Campaign, "shipment": Shipment,
        "shipment_item": ShipmentItem, "cost_item": CostItem, "deliverable": Deliverable,
        "project": Project, "product": Product, "project_product": ProjectProduct,
    }
    actions = json.loads(batch.undo_json or "[]")
    removed = restored = 0
    for action in reversed(actions):
        model = model_map.get(action.get("entity"))
        if not model:
            continue
        entity = db.get(model, action.get("id"))
        if action.get("kind") == "create" and entity:
            db.delete(entity)
            db.flush()
            removed += 1
        elif action.get("kind") == "update" and entity:
            for field, old_value in (action.get("before") or {}).items():
                setattr(entity, field, old_value)
            restored += 1
    batch.status = "undone"
    batch.undone_at = datetime.utcnow()
    add_audit_log(db, user, "undo", "import_batch", batch.id, after={"removed": removed, "restored": restored}, reason=change_reason)
    db.commit()
    return {"ok": True, "removed": removed, "restored": restored}


@app.post("/api/import/preview")
async def import_preview(db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))], file: UploadFile = File(...)):
    result = preview_import(await file.read(), db)
    return result.__dict__


@app.post("/api/import/confirm")
async def import_confirm(db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))], file: UploadFile = File(...)):
    result = confirm_import(db, await file.read(), file.filename, user.id)
    return result.__dict__


@app.post("/api/execution-import/preview")
async def execution_import_preview(db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))], file: UploadFile = File(...)):
    return preview_execution_import(await file.read(), db)


@app.post("/api/execution-import/confirm")
async def execution_import_confirm(db: Annotated[Session, Depends(get_db)], user: Annotated[User, Depends(require_roles("Admin"))], file: UploadFile = File(...)):
    return confirm_execution_import(db, await file.read(), file.filename, user.id)


if FRONTEND_DIST.exists():
    app.mount("/assets", StaticFiles(directory=FRONTEND_DIST / "assets"), name="assets")


@app.get("/{path:path}")
def spa(path: str):
    index = FRONTEND_DIST / "index.html"
    if index.exists():
        return FileResponse(index)
    return {"message": "Frontend is not built yet. Run start.bat or build the frontend."}
