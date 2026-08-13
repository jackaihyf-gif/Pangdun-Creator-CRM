from __future__ import annotations

from dataclasses import dataclass
import hashlib
from io import BytesIO
import json
from typing import Any

import openpyxl
from sqlalchemy.orm import Session

from .models import Campaign, CampaignStageEvent, CostItem, Deliverable, ImportBatch, Media, Product, Project, ProjectProduct, Shipment, ShipmentItem
from .media_taxonomy import normalize_media_payload
from .profile_links import clean_profile_links, profile_identity
from .product_backfill import ensure_project_link, find_or_create_product, find_product_matches


STATUS_MAP = {
    "已产出": "已发布",
    "已到货待产出": "已签收待产出",
    "已发货待收": "运输中",
    "代理发货": "运输中",
    "待发货": "待发货",
}


def value(row: dict[str, Any], key: str) -> str | None:
    item = row.get(key)
    return str(item).strip() if item not in (None, "") else None


def split_products(value_text: str | None) -> list[str]:
    return [item.strip() for item in (value_text or "").splitlines() if item.strip()]


def number(row: dict[str, Any], key: str) -> float | None:
    item = row.get(key)
    if item in (None, ""):
        return None
    try:
        return float(item)
    except (TypeError, ValueError):
        return None


def load_execution_rows(content: bytes) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    values = sheet.iter_rows(values_only=True)
    headers = [str(cell).strip() if cell else "" for cell in next(values, [])]
    rows = []
    for row_number, row in enumerate(values, start=2):
        item = dict(zip(headers, row))
        if any(cell not in (None, "") for cell in row):
            item["_row_number"] = row_number
            rows.append(item)
    return rows


def preview_execution_import(content: bytes, db: Session | None = None) -> dict[str, Any]:
    rows = load_execution_rows(content)
    previews = []
    warning_count = 0
    for row in rows:
        code = value(row, "OA PI编号")
        channel = value(row, "Channel") or "未命名渠道"
        warnings = []
        if not code:
            warnings.append("缺少 OA/PI，将进入历史导入待归类项目")
        if not value(row, "频道链接"):
            warnings.append("缺少频道链接，媒体去重需人工确认")
        if db and value(row, "频道链接") and len(find_media_identity_matches(db, value(row, "频道链接"))) > 1:
            warnings.append("频道主页匹配到多个媒体，需先合并")
        if db:
            for raw_product in split_products(value(row, "产品类型")):
                if len(find_product_matches(db, raw_product)) > 1:
                    warnings.append(f"产品“{raw_product}”匹配到多个型号或别名")
        warning_count += len(warnings)
        action = "新增"
        if db:
            project = db.query(Project).filter(Project.project_code == (code or "HISTORY-UNSORTED")).first()
            media = find_or_create_media(db, row, create=False)
            campaign = find_existing_campaign(db, project, media, row) if project and media else None
            if campaign:
                action = "更新" if campaign_needs_update(campaign, row) else "跳过"
        previews.append({
            "row_number": row["_row_number"],
            "project_code": code,
            "media_name": channel,
            "country": value(row, "国家"),
            "channel": value(row, "渠道"),
            "execution_status": STATUS_MAP.get(value(row, "进度") or "", "待确认"),
            "product_bundle": value(row, "产品类型"),
            "tracking_number": value(row, "追踪编号"),
            "content_url": value(row, "产出内容链接"),
            "warnings": warnings,
            "import_action": action,
        })
    return {
        "total": len(rows), "warning_count": warning_count, "rows": previews,
        "created_count": sum(1 for item in previews if item["import_action"] == "新增"),
        "updated_count": sum(1 for item in previews if item["import_action"] == "更新"),
        "unchanged_count": sum(1 for item in previews if item["import_action"] == "跳过"),
        "conflict_count": sum(1 for item in previews if item["warnings"]),
    }


def find_media_identity_matches(db: Session, url: str | None) -> list[Media]:
    identity = profile_identity(url)
    if not identity:
        return []
    return [item for item in db.query(Media).all() if identity in {profile_identity(link.get("url")) for link in clean_profile_links(item.profile_links, item.website_url)}]


def find_or_create_media(db: Session, row: dict[str, Any], create: bool = True) -> Media | None:
    name = value(row, "Channel") or "未命名渠道"
    url = value(row, "频道链接")
    if url:
        matches = find_media_identity_matches(db, url)
        if len(matches) > 1:
            raise ValueError("频道主页匹配到多个媒体，请先合并后再导入")
        item = matches[0] if matches else None
    else:
        normalized = normalize_media_payload({"name": name, "country": value(row, "国家"), "platform_type": value(row, "渠道")})
        item = db.query(Media).filter(Media.name == name, Media.country == normalized.get("country")).first()
    if item:
        return item
    if not create:
        return None
    data = normalize_media_payload({"name": name, "country": value(row, "国家"), "platform_type": value(row, "渠道"), "website_url": url})
    data["profile_links"] = clean_profile_links(None, url)
    data["website_url"] = data["profile_links"][0]["url"] if data["profile_links"] else None
    item = Media(**data)
    db.add(item)
    db.flush()
    return item


def find_existing_campaign(db: Session, project: Project | None, media: Media | None, row: dict[str, Any]) -> Campaign | None:
    if not project or not media:
        return None
    query = db.query(Campaign).filter(Campaign.project_id == project.id, Campaign.media_id == media.id)
    tracking = value(row, "追踪编号")
    if tracking:
        return query.filter(Campaign.shipments.any(Shipment.tracking_number == tracking)).first()
    content_url = value(row, "产出内容链接")
    if content_url:
        return query.filter(Campaign.deliverables.any(Deliverable.url == content_url)).first()
    notes = "\n".join(part for part in [value(row, "合作备注"), value(row, "合作备注2（当前情况）")] if part)
    return query.filter(Campaign.collaboration_type == value(row, "推广形式"), Campaign.notes == (notes or None)).first()


def campaign_needs_update(campaign: Campaign, row: dict[str, Any]) -> bool:
    expected = {
        "collaboration_type": value(row, "推广形式"),
        "execution_status": STATUS_MAP.get(value(row, "进度") or "", "待确认"),
        "stage": "Published" if value(row, "进度") == "已产出" else "Not Started",
    }
    return any(incoming not in (None, "") and getattr(campaign, field) != incoming for field, incoming in expected.items())


def set_with_undo(item: Any, entity: str, values: dict[str, Any], actions: list[dict[str, Any]]) -> bool:
    before = {}
    for field, incoming in values.items():
        if incoming not in (None, "") and getattr(item, field) != incoming:
            before[field] = getattr(item, field)
            setattr(item, field, incoming)
    if before:
        actions.append({"kind": "update", "entity": entity, "id": item.id, "before": before})
    return bool(before)


def confirm_execution_import(db: Session, content: bytes, filename: str | None = None, user_id: int | None = None) -> dict[str, Any]:
    rows = load_execution_rows(content)
    actions: list[dict[str, Any]] = []
    batch = ImportBatch(import_type="execution", filename=filename, source_hash=hashlib.sha256(content).hexdigest(), user_id=user_id, status="processing")
    db.add(batch)
    db.flush()
    fallback = db.query(Project).filter(Project.project_code == "HISTORY-UNSORTED").first()
    if not fallback:
        fallback = Project(name="历史导入待归类", project_code="HISTORY-UNSORTED", status="Active", notes="来自费用统计表、缺少 OA/PI 编号的历史记录")
        db.add(fallback)
        db.flush()
        actions.append({"kind": "create", "entity": "project", "id": fallback.id})
    imported = created = updated = unchanged = 0
    projects: dict[str, Project] = {}
    for row in rows:
        code = value(row, "OA PI编号")
        project = fallback
        if code:
            project = projects.get(code) or db.query(Project).filter(Project.project_code == code).first()
            if not project:
                project = Project(name=f"历史导入 {code}", project_code=code, status="Active")
                db.add(project)
                db.flush()
                actions.append({"kind": "create", "entity": "project", "id": project.id})
            projects[code] = project
        media = find_or_create_media(db, row, create=False)
        if not media:
            media = find_or_create_media(db, row)
            actions.append({"kind": "create", "entity": "media", "id": media.id})
        campaign = find_existing_campaign(db, project, media, row)
        notes = "\n".join(part for part in [value(row, "合作备注"), value(row, "合作备注2（当前情况）")] if part)
        campaign_values = {
            "collaboration_type": value(row, "推广形式"), "execution_status": STATUS_MAP.get(value(row, "进度") or "", "待确认"),
            "stage": "Published" if value(row, "进度") == "已产出" else "Not Started", "notes": notes or None,
        }
        row_changed = False
        if not campaign:
            campaign = Campaign(project_id=project.id, media_id=media.id, **campaign_values)
            db.add(campaign)
            db.flush()
            db.add(CampaignStageEvent(campaign_id=campaign.id, user_id=user_id, from_status=None, to_status=campaign.execution_status, action="import", reason="历史执行表导入"))
            actions.append({"kind": "create", "entity": "campaign", "id": campaign.id})
            created += 1
            row_changed = True
        else:
            row_changed = set_with_undo(campaign, "campaign", campaign_values, actions)
        shipment = campaign.shipments[0] if campaign.shipments else None
        shipment_values = {"recipient_address": value(row, "地址信息"), "oa_pi_number": code, "tracking_number": value(row, "追踪编号"), "status": STATUS_MAP.get(value(row, "进度") or "", "待确认")}
        if not shipment:
            shipment = Shipment(campaign_id=campaign.id, **shipment_values)
            db.add(shipment)
            db.flush()
            actions.append({"kind": "create", "entity": "shipment", "id": shipment.id})
            row_changed = True
        else:
            row_changed = set_with_undo(shipment, "shipment", shipment_values, actions) or row_changed
        product_bundle = value(row, "产品类型")
        if product_bundle:
            for product_name in split_products(product_bundle):
                if product_name.strip():
                    raw_product = product_name.strip()
                    product = db.query(Product).filter(Product.model == raw_product).first()
                    if not product:
                        product = find_or_create_product(db, raw_product, "费用表导入")
                        actions.append({"kind": "create", "entity": "product", "id": product.id})
                    existing_link = db.query(ProjectProduct).filter(ProjectProduct.project_id == project.id, ProjectProduct.product_id == product.id).first()
                    ensure_project_link(db, project.id, product.id)
                    if not existing_link:
                        link = db.query(ProjectProduct).filter(ProjectProduct.project_id == project.id, ProjectProduct.product_id == product.id).first()
                        actions.append({"kind": "create", "entity": "project_product", "id": link.id})
                    shipment_item = db.query(ShipmentItem).filter(ShipmentItem.shipment_id == shipment.id, ShipmentItem.product_id == product.id).first()
                    if not shipment_item:
                        shipment_item = ShipmentItem(shipment_id=shipment.id, product_id=product.id, product_name=product.model)
                        db.add(shipment_item)
                        db.flush()
                        actions.append({"kind": "create", "entity": "shipment_item", "id": shipment_item.id})
                        row_changed = True
        for label, source in [("产品费用", "产品费用"), ("物流/关税", "运费/保费/关税预付"), ("评测费用", "评测费用")]:
            amount = number(row, source)
            if amount is not None:
                cost = db.query(CostItem).filter(CostItem.campaign_id == campaign.id, CostItem.cost_type == label).first()
                if not cost:
                    cost = CostItem(campaign_id=campaign.id, cost_type=label, actual_amount=amount, currency="CNY", payment_status="已付款")
                    db.add(cost)
                    db.flush()
                    actions.append({"kind": "create", "entity": "cost_item", "id": cost.id})
                    row_changed = True
                else:
                    row_changed = set_with_undo(cost, "cost_item", {"actual_amount": amount, "payment_status": "已付款"}, actions) or row_changed
        content_url = value(row, "产出内容链接")
        if content_url:
            deliverable = db.query(Deliverable).filter(Deliverable.campaign_id == campaign.id, Deliverable.url == content_url).first()
            if not deliverable:
                deliverable = Deliverable(campaign_id=campaign.id, deliverable_type=value(row, "渠道") or "Other", url=content_url, views=None)
                db.add(deliverable)
                db.flush()
                actions.append({"kind": "create", "entity": "deliverable", "id": deliverable.id})
                row_changed = True
        if campaign.id and not row_changed:
            unchanged += 1
        elif campaign.id and campaign not in db.new and not any(action.get("entity") == "campaign" and action.get("id") == campaign.id and action.get("kind") == "create" for action in actions):
            updated += 1
        imported += 1
    summary = {"success_count": imported, "created_count": created, "updated_count": updated, "unchanged_count": unchanged, "conflict_count": sum(1 for row in rows if not value(row, "OA PI编号") or not value(row, "频道链接")), "project_count": len(projects) + 1, "fallback_count": sum(1 for row in rows if not value(row, "OA PI编号"))}
    batch.status = "completed"
    batch.summary_json = json.dumps(summary, ensure_ascii=False)
    batch.undo_json = json.dumps(actions, ensure_ascii=False)
    db.commit()
    return {**summary, "batch_id": batch.id}
